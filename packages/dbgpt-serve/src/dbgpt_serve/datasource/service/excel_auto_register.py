#!/usr/bin/env python3
"""
Excel 自动注册到数据源服务
支持自动缓存和增量导入
"""
# ruff: noqa: E501

import hashlib
import json
import logging
import os
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
import pandas as pd

logger = logging.getLogger(__name__)


class ExcelCacheManager:
    """Excel 缓存管理器"""

    def __init__(self, cache_dir: str = None):
        """
        初始化缓存管理器

        Args:
            cache_dir: 缓存目录，默认为 pilot/data/excel_cache
        """
        if cache_dir is None:
            # 使用相对路径
            current_dir = Path(__file__).parent
            cache_dir = (
                current_dir.parent.parent.parent.parent.parent
                / "pilot"
                / "data"
                / "excel_cache"
            )

        self.cache_dir = Path(cache_dir)
        self.cache_dir.mkdir(parents=True, exist_ok=True)

        # 元数据数据库路径
        self.meta_db_path = self.cache_dir / "excel_metadata.db"
        self._init_metadata_db()

    def _init_metadata_db(self):
        """初始化元数据数据库"""
        conn = sqlite3.connect(str(self.meta_db_path))
        cursor = conn.cursor()

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS excel_metadata (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                content_hash VARCHAR(64) UNIQUE NOT NULL,
                original_filename VARCHAR(255) NOT NULL,
                table_name VARCHAR(255) NOT NULL,
                db_name VARCHAR(255) NOT NULL,
                db_path TEXT NOT NULL,
                row_count INTEGER NOT NULL,
                column_count INTEGER NOT NULL,
                columns_info TEXT NOT NULL,
                summary_prompt TEXT,
                data_schema_json TEXT,
                id_columns TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                last_accessed TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                access_count INTEGER DEFAULT 0
            )
        """)
        
        # 新增：多表元数据表（一个Excel文件可以有多个sheet/表）
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS excel_tables_metadata (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                file_hash VARCHAR(64) NOT NULL,
                sheet_name VARCHAR(255) NOT NULL,
                table_hash VARCHAR(64) UNIQUE NOT NULL,
                original_filename VARCHAR(255) NOT NULL,
                table_name VARCHAR(255) NOT NULL,
                db_name VARCHAR(255) NOT NULL,
                db_path TEXT NOT NULL,
                row_count INTEGER NOT NULL,
                column_count INTEGER NOT NULL,
                columns_info TEXT NOT NULL,
                summary_prompt TEXT,
                data_schema_json TEXT,
                id_columns TEXT,
                create_table_sql TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                last_accessed TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                access_count INTEGER DEFAULT 0,
                UNIQUE(file_hash, sheet_name)
            )
        """)

        try:
            cursor.execute("SELECT data_schema_json FROM excel_metadata LIMIT 1")
        except sqlite3.OperationalError:
            cursor.execute(
                "ALTER TABLE excel_metadata ADD COLUMN data_schema_json TEXT"
            )
        
        try:
            cursor.execute("SELECT id_columns FROM excel_metadata LIMIT 1")
        except sqlite3.OperationalError:
            cursor.execute(
                "ALTER TABLE excel_metadata ADD COLUMN id_columns TEXT"
            )
        
        # 为新表添加可能缺失的列
        try:
            cursor.execute("SELECT create_table_sql FROM excel_tables_metadata LIMIT 1")
        except sqlite3.OperationalError:
            cursor.execute(
                "ALTER TABLE excel_tables_metadata ADD COLUMN create_table_sql TEXT"
            )

        conn.commit()
        conn.close()

    @staticmethod
    def calculate_excel_hash(df: pd.DataFrame, filename: str) -> str:
        """
        计算 Excel 的内容哈希值（基于DataFrame，用于向后兼容）

        Args:
            df: DataFrame 对象
            filename: 文件名（包含在哈希计算中）

        Returns:
            SHA256 哈希值
        """
        # 组合文件名、列名和数据内容
        content_parts = [
            filename,
            ",".join(df.columns.tolist()),
            df.to_csv(index=False),
        ]
        content = "\n".join(content_parts)

        return hashlib.sha256(content.encode("utf-8")).hexdigest()

    @staticmethod
    def calculate_file_hash(file_path: str, sheet_names: List[str] = None) -> str:
        """
        计算文件级别的哈希值（基于文件内容和sheet列表）

        Args:
            file_path: Excel文件路径
            sheet_names: 要处理的sheet名称列表（如果为None则不考虑sheet信息）

        Returns:
            SHA256 哈希值
        """
        sha256_hash = hashlib.sha256()

        # 读取文件内容并计算哈希
        with open(file_path, "rb") as f:
            # 分块读取，避免大文件占用过多内存
            for byte_block in iter(lambda: f.read(4096), b""):
                sha256_hash.update(byte_block)

        # 如果指定了sheet_names，将其也纳入哈希计算（确保不同sheet组合产生不同哈希）
        if sheet_names:
            sheet_info = ",".join(sorted(sheet_names))
            sha256_hash.update(sheet_info.encode("utf-8"))

        return sha256_hash.hexdigest()

    def get_cached_info(self, content_hash: str) -> Optional[Dict]:
        """根据内容哈希获取缓存信息"""
        conn = sqlite3.connect(str(self.meta_db_path))
        cursor = conn.cursor()
        
        cursor.execute(
            """
            SELECT 
                content_hash, original_filename, table_name, db_name, db_path,
                row_count, column_count, columns_info, summary_prompt, data_schema_json,
                id_columns, created_at, last_accessed, access_count
            FROM excel_metadata
            WHERE content_hash = ?
        """,
            (content_hash,),
        )
        row = cursor.fetchone()

        if row:
            cursor.execute(
                """
                UPDATE excel_metadata
                SET last_accessed = CURRENT_TIMESTAMP, access_count = access_count + 1
                WHERE content_hash = ?
            """,
                (content_hash,),
            )
            conn.commit()
            conn.close()
            return {
                "content_hash": row[0],
                "original_filename": row[1],
                "table_name": row[2],
                "db_name": row[3],
                "db_path": row[4],
                "row_count": row[5],
                "column_count": row[6],
                "columns_info": json.loads(row[7]),
                "summary_prompt": row[8],
                "data_schema_json": row[9],
                "id_columns": json.loads(row[10]) if row[10] else [],
                "created_at": row[11],
                "last_accessed": row[12],
                "access_count": row[13],
            }

        conn.close()
        return None

    def save_cache_info(
        self,
        content_hash: str,
        original_filename: str,
        table_name: str,
        db_name: str,
        db_path: str,
        df: pd.DataFrame,
        summary_prompt: str = None,
        data_schema_json: str = None,
        id_columns: List[str] = None,
    ):
        """
        保存缓存信息

        Args:
            content_hash: 内容哈希值
            original_filename: 原始文件名
            table_name: 表名
            db_name: 数据库名
            db_path: 数据库路径
            df: DataFrame 对象
            summary_prompt: 数据理解提示词
            data_schema_json: 数据schema JSON
            id_columns: ID列名列表
        """
        conn = sqlite3.connect(str(self.meta_db_path))
        cursor = conn.cursor()

        columns_info = [
            {"name": col, "dtype": str(df[col].dtype)} for col in df.columns
        ]

        cursor.execute(
            """
            INSERT OR REPLACE INTO excel_metadata
            (content_hash, original_filename, table_name, db_name, db_path,
             row_count, column_count, columns_info, summary_prompt, data_schema_json, id_columns)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
            (
                content_hash,
                original_filename,
                table_name,
                db_name,
                db_path,
                len(df),
                len(df.columns),
                json.dumps(columns_info, ensure_ascii=False),
                summary_prompt,
                data_schema_json,
                json.dumps(id_columns if id_columns else [], ensure_ascii=False),
            ),
        )

        conn.commit()
        conn.close()

    def update_summary_prompt(self, content_hash: str, summary_prompt: str):
        """
        更新数据理解提示词

        Args:
            content_hash: 内容哈希值
            summary_prompt: 数据理解提示词
        """
        conn = sqlite3.connect(str(self.meta_db_path))
        cursor = conn.cursor()

        cursor.execute(
            """
            UPDATE excel_metadata
            SET summary_prompt = ?
            WHERE content_hash = ?
        """,
            (summary_prompt, content_hash),
        )

        conn.commit()
        conn.close()

    def delete_cache_by_hash(self, content_hash: str):
        """
        根据内容哈希删除缓存

        Args:
            content_hash: 内容哈希值
        """
        conn = sqlite3.connect(str(self.meta_db_path))
        cursor = conn.cursor()

        cursor.execute(
            "DELETE FROM excel_metadata WHERE content_hash = ?", (content_hash,)
        )
        deleted = cursor.rowcount

        conn.commit()
        conn.close()

        return deleted > 0

    def delete_cache_by_filename(self, filename: str):
        """
        根据文件名删除缓存

        Args:
            filename: 原始文件名
        """
        conn = sqlite3.connect(str(self.meta_db_path))
        cursor = conn.cursor()

        cursor.execute(
            "DELETE FROM excel_metadata WHERE original_filename = ?", (filename,)
        )
        deleted = cursor.rowcount

        conn.commit()
        conn.close()

        return deleted > 0

    def list_all_cache(self) -> List[Dict]:
        """
        列出所有缓存记录

        Returns:
            缓存记录列表
        """
        conn = sqlite3.connect(str(self.meta_db_path))
        cursor = conn.cursor()

        cursor.execute("""
            SELECT 
                content_hash, original_filename, table_name, db_name,
                row_count, column_count, created_at, last_accessed, access_count
            FROM excel_metadata
            ORDER BY last_accessed DESC
        """)

        records = cursor.fetchall()
        conn.close()

        result = []
        for row in records:
            result.append(
                {
                    "content_hash": row[0],
                    "original_filename": row[1],
                    "table_name": row[2],
                    "db_name": row[3],
                    "row_count": row[4],
                    "column_count": row[5],
                    "created_at": row[6],
                    "last_accessed": row[7],
                    "access_count": row[8],
                }
            )

        return result

    def get_tables_by_file_hash(self, file_hash: str) -> List[Dict]:
        """
        根据文件哈希获取所有表的缓存信息（多表模式）
        
        Args:
            file_hash: 文件哈希值
            
        Returns:
            表信息列表
        """
        conn = sqlite3.connect(str(self.meta_db_path))
        cursor = conn.cursor()
        
        cursor.execute(
            """
            SELECT 
                file_hash, sheet_name, table_hash, original_filename, table_name,
                db_name, db_path, row_count, column_count, columns_info,
                summary_prompt, data_schema_json, id_columns, create_table_sql,
                created_at, last_accessed, access_count
            FROM excel_tables_metadata
            WHERE file_hash = ?
            ORDER BY id
        """,
            (file_hash,),
        )
        rows = cursor.fetchall()
        
        if rows:
            # 更新访问时间和次数
            cursor.execute(
                """
                UPDATE excel_tables_metadata
                SET last_accessed = CURRENT_TIMESTAMP, access_count = access_count + 1
                WHERE file_hash = ?
            """,
                (file_hash,),
            )
            conn.commit()
        
        conn.close()
        
        result = []
        for row in rows:
            result.append({
                "file_hash": row[0],
                "sheet_name": row[1],
                "table_hash": row[2],
                "original_filename": row[3],
                "table_name": row[4],
                "db_name": row[5],
                "db_path": row[6],
                "row_count": row[7],
                "column_count": row[8],
                "columns_info": json.loads(row[9]) if row[9] else [],
                "summary_prompt": row[10],
                "data_schema_json": row[11],
                "id_columns": json.loads(row[12]) if row[12] else [],
                "create_table_sql": row[13],
                "created_at": row[14],
                "last_accessed": row[15],
                "access_count": row[16],
            })
        
        return result

    def save_table_cache_info(
        self,
        file_hash: str,
        sheet_name: str,
        table_hash: str,
        original_filename: str,
        table_name: str,
        db_name: str,
        db_path: str,
        df: pd.DataFrame,
        summary_prompt: str = None,
        data_schema_json: str = None,
        id_columns: List[str] = None,
        create_table_sql: str = None,
    ):
        """
        保存单个表的缓存信息（多表模式）
        
        Args:
            file_hash: 文件哈希值
            sheet_name: sheet名称
            table_hash: 表哈希值（file_hash + sheet_name的组合哈希）
            original_filename: 原始文件名
            table_name: 表名
            db_name: 数据库名
            db_path: 数据库路径
            df: DataFrame 对象
            summary_prompt: 数据理解提示词
            data_schema_json: 数据schema JSON
            id_columns: ID列名列表
            create_table_sql: 建表SQL语句
        """
        conn = sqlite3.connect(str(self.meta_db_path))
        cursor = conn.cursor()

        columns_info = [
            {"name": col, "dtype": str(df[col].dtype)} for col in df.columns
        ]

        cursor.execute(
            """
            INSERT OR REPLACE INTO excel_tables_metadata
            (file_hash, sheet_name, table_hash, original_filename, table_name, 
             db_name, db_path, row_count, column_count, columns_info, 
             summary_prompt, data_schema_json, id_columns, create_table_sql)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
            (
                file_hash,
                sheet_name,
                table_hash,
                original_filename,
                table_name,
                db_name,
                db_path,
                len(df),
                len(df.columns),
                json.dumps(columns_info, ensure_ascii=False),
                summary_prompt,
                data_schema_json,
                json.dumps(id_columns if id_columns else [], ensure_ascii=False),
                create_table_sql,
            ),
        )

        conn.commit()
        conn.close()

    def delete_tables_by_file_hash(self, file_hash: str) -> int:
        """
        根据文件哈希删除所有相关表的缓存
        
        Args:
            file_hash: 文件哈希值
            
        Returns:
            删除的记录数
        """
        conn = sqlite3.connect(str(self.meta_db_path))
        cursor = conn.cursor()

        cursor.execute(
            "DELETE FROM excel_tables_metadata WHERE file_hash = ?", (file_hash,)
        )
        deleted = cursor.rowcount

        conn.commit()
        conn.close()

        return deleted


class ExcelAutoRegisterService:
    """Excel 自动注册到数据源服务"""

    _instance = None
    _lock = None

    def __new__(cls, *args, **kwargs):
        """单例模式：确保只有一个实例"""
        if cls._instance is None:
            # 创建锁（如果还没有）
            if cls._lock is None:
                import threading

                cls._lock = threading.Lock()

            with cls._lock:
                # 双重检查
                if cls._instance is None:
                    cls._instance = super().__new__(cls)
        return cls._instance

    def __init__(self, llm_client=None, model_name=None, system_app=None):
        """初始化服务"""
        if not hasattr(self, "_initialized"):
            self.cache_manager = ExcelCacheManager()
            self.llm_client = llm_client
            self.model_name = model_name
            self._system_app = system_app

            current_dir = Path(__file__).parent
            base_dir = (
                current_dir.parent.parent.parent.parent.parent / "pilot" / "meta_data"
            )
            self.db_storage_dir = base_dir / "excel_dbs"
            self.db_storage_dir.mkdir(parents=True, exist_ok=True)

            self._initialized = True
        else:
            # 每次都更新 llm_client 和 model_name，避免使用缓存的旧值
            self.llm_client = llm_client
            self.model_name = model_name
            if system_app is not None:
                self._system_app = system_app

    def _get_llm_client_and_model(self):
        """获取 LLM 客户端和模型名称"""
        default_model = None
        try:
            from dbgpt._private.config import Config
            cfg = Config()
            default_model = cfg.LLM_MODEL
        except Exception:
            pass
        
        if hasattr(self, '_system_app') and self._system_app is not None:
            try:
                from dbgpt.component import ComponentType
                from dbgpt.model.cluster import WorkerManagerFactory
                from dbgpt.model.cluster.client import DefaultLLMClient
                
                worker_manager = self._system_app.get_component(
                    ComponentType.WORKER_MANAGER_FACTORY, WorkerManagerFactory
                ).create()
                
                if default_model:
                    return DefaultLLMClient(worker_manager, auto_convert_message=True), default_model
            except Exception as e:
                logger.warning(f"创建LLM客户端失败: {e}")
        
        return None, default_model
    
    def _get_default_model_name(self) -> Optional[str]:
        """获取默认的 LLM 模型名称"""
        _, model_name = self._get_llm_client_and_model()
        return model_name

    def _remove_excel_filters(self, excel_file_path: str) -> str:
        """去除 Excel 文件的筛选状态"""
        import zipfile
        import tempfile
        import shutil
        import re
        
        file_ext = Path(excel_file_path).suffix.lower()
        if file_ext == '.xls':
            return excel_file_path
        
        try:
            temp_dir = tempfile.mkdtemp()
            temp_xlsx = os.path.join(temp_dir, "temp_output.xlsx")
            filters_removed = False
            
            with zipfile.ZipFile(excel_file_path, 'r') as zip_in:
                with zipfile.ZipFile(temp_xlsx, 'w', zipfile.ZIP_DEFLATED) as zip_out:
                    for item in zip_in.namelist():
                        data = zip_in.read(item)
                        
                        if item.startswith('xl/worksheets/sheet') and item.endswith('.xml'):
                            content = data.decode('utf-8')
                            original_content = content
                            
                            # 删除 autoFilter 和 hidden 属性
                            content = re.sub(r'<autoFilter[^>]*/>', '', content)
                            content = re.sub(r'<autoFilter[^>]*>.*?</autoFilter>', '', content, flags=re.DOTALL)
                            content = re.sub(r'(<row[^>]*)\s+hidden="1"([^>]*>)', r'\1\2', content)
                            content = re.sub(r'(<row[^>]*)\s+hidden="true"([^>]*>)', r'\1\2', content, flags=re.IGNORECASE)
                            
                            if content != original_content:
                                filters_removed = True
                            data = content.encode('utf-8')
                        
                        zip_out.writestr(item, data)
            
            if filters_removed:
                shutil.move(temp_xlsx, excel_file_path)
            else:
                os.remove(temp_xlsx)
            shutil.rmtree(temp_dir, ignore_errors=True)
            return excel_file_path
            
        except Exception as e:
            logger.warning(f"去除筛选失败，尝试openpyxl: {e}")
            return self._remove_excel_filters_openpyxl(excel_file_path)
    
    def _remove_excel_filters_openpyxl(self, excel_file_path: str) -> str:
        """使用 openpyxl 去除筛选状态（备用方案）"""
        try:
            wb = openpyxl.load_workbook(excel_file_path)
            filters_removed = False
            
            for ws in wb.worksheets:
                if ws.auto_filter and ws.auto_filter.ref:
                    ws.auto_filter.ref = None
                    filters_removed = True
                for row in ws.row_dimensions:
                    if ws.row_dimensions[row].hidden:
                        ws.row_dimensions[row].hidden = False
                        filters_removed = True
            
            if filters_removed:
                wb.save(excel_file_path)
            wb.close()
            return excel_file_path
        except Exception as e:
            logger.warning(f"openpyxl去除筛选失败: {e}")
            return excel_file_path

    def _is_csv_file(self, file_path: str) -> bool:
        """判断是否为CSV文件"""
        file_ext = Path(file_path).suffix.lower()
        return file_ext in ['.csv', '.txt', '.tsv']
    
    def _read_csv_sample(self, csv_file_path: str, nrows: int = 1000) -> pd.DataFrame:
        """
        读取CSV文件的样本数据（用于表头分析和列类型推断）
        
        Args:
            csv_file_path: CSV文件路径
            nrows: 读取的行数
            
        Returns:
            DataFrame
        """
        try:
            # 尝试自动检测分隔符和编码
            encodings = ['utf-8', 'gbk', 'gb2312', 'gb18030', 'latin1']
            
            for encoding in encodings:
                try:
                    df = pd.read_csv(
                        csv_file_path,
                        nrows=nrows,
                        encoding=encoding,
                        low_memory=False
                    )
                    logger.info(f"CSV文件编码: {encoding}, 列数: {len(df.columns)}, 样本行数: {len(df)}")
                    return df
                except UnicodeDecodeError:
                    continue
                except Exception as e:
                    logger.warning(f"使用编码 {encoding} 读取CSV失败: {e}")
                    continue
            
            raise Exception("无法使用任何编码读取CSV文件")
        except Exception as e:
            logger.error(f"读取CSV样本失败: {e}")
            raise

    def _read_excel_file(self, excel_file_path: str, sheet_name=None, header=None) -> pd.DataFrame:
        """
        智能读取 Excel 文件，支持 .xls 和 .xlsx 格式
        
        Args:
            excel_file_path: Excel 文件路径
            sheet_name: sheet 名称
            header: 表头行索引
            
        Returns:
            DataFrame
        """
        file_ext = Path(excel_file_path).suffix.lower()
        
        try:
            if file_ext == '.xls':
                # 使用 xlrd 引擎读取旧版 .xls 文件
                logger.info(f"检测到 .xls 格式，使用 xlrd 引擎读取")
                return pd.read_excel(
                    excel_file_path, 
                    sheet_name=sheet_name, 
                    header=header,
                    engine='xlrd'
                )
            else:
                # 使用默认的 openpyxl 引擎读取 .xlsx 文件
                return pd.read_excel(
                    excel_file_path, 
                    sheet_name=sheet_name, 
                    header=header
                )
        except Exception as e:
            logger.error(f"读取 Excel 文件失败: {e}")
            # 如果默认方式失败，尝试另一种引擎
            try:
                if file_ext == '.xls':
                    logger.warning(f"xlrd 读取失败，尝试使用 openpyxl")
                    return pd.read_excel(
                        excel_file_path, 
                        sheet_name=sheet_name, 
                        header=header,
                        engine='openpyxl'
                    )
                else:
                    logger.warning(f"openpyxl 读取失败，尝试使用 xlrd")
                    return pd.read_excel(
                        excel_file_path, 
                        sheet_name=sheet_name, 
                        header=header,
                        engine='xlrd'
                    )
            except Exception as e2:
                logger.error(f"所有引擎都无法读取文件: {e2}")
                raise Exception(f"无法读取 Excel 文件 {excel_file_path}: {e2}")

    def _get_cell_value(self, cell) -> Optional[str]:
        """获取单元格值，处理公式"""
        if cell.value is None:
            return None

        if cell.data_type == "f":
            try:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cleaned = self._clean_excel_formula(cell.value)
                    return cleaned if cleaned else None
            except Exception as e:
                logger.warning(f"获取公式计算结果失败: {e}")

        value_str = str(cell.value)
        return (
            value_str.replace("\n", "")
            .replace("\r", "")
            .replace("\t", "")
            .replace(" ", "")
        )

    def _get_cell_bg_color(self, cell) -> Optional[str]:
        """获取单元格背景色"""
        fill = cell.fill
        if fill.fill_type == "solid" or fill.fill_type == "patternFill":
            fg_color = fill.fgColor
            if fg_color.type == "rgb":
                rgb = fg_color.rgb
                if rgb and len(rgb) == 8:
                    return rgb[2:]
                return rgb
            elif fg_color.type == "indexed":
                return f"indexed_{fg_color.indexed}"
            elif fg_color.type == "theme":
                tint = fg_color.tint if fg_color.tint else 0
                return f"theme_{fg_color.theme}_{tint:.2f}"
        return None

    def _detect_header_rows_with_color(
        self, excel_file_path: str, sheet_name: str = None
    ) -> Tuple[List[int], Dict]:
        """使用颜色信息和LLM检测表头行

        Args:
            excel_file_path: Excel文件路径
            sheet_name: sheet名称，如果为None则使用active sheet
        """
        file_ext = Path(excel_file_path).suffix.lower()
        
        rows_data = []
        rows_colors = []
        max_cols = 0
        
        # .xls 文件不支持 openpyxl，跳过颜色检测，其他流程保持一致
        if file_ext == '.xls':
            logger.info("检测到 .xls 格式，跳过颜色检测")
            # 读取前几行数据
            df_preview = self._read_excel_file(excel_file_path, sheet_name=sheet_name, header=None)
            
            max_check_rows = min(20, len(df_preview))
            max_cols = len(df_preview.columns) if len(df_preview) > 0 else 0
            
            for i in range(max_check_rows):
                row_values = [str(v) if pd.notna(v) else "" for v in df_preview.iloc[i].tolist()]
                row_colors = [None] * len(row_values)  # .xls 不支持颜色检测，全部设为 None
                rows_data.append(row_values)
                rows_colors.append(row_colors)
        else:
            # .xlsx 文件尝试使用 openpyxl 读取颜色信息
            try:
                wb = openpyxl.load_workbook(excel_file_path)
                ws = wb[sheet_name] if sheet_name else wb.active

                max_check_rows = min(20, ws.max_row)
                max_cols = ws.max_column

                for row_idx in range(1, max_check_rows + 1):
                    row_values = []
                    row_colors = []
                    for col_idx in range(1, max_cols + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell_value = self._get_cell_value(cell)
                        row_values.append(cell_value if cell_value is not None else "")
                        row_colors.append(self._get_cell_bg_color(cell))
                    rows_data.append(row_values)
                    rows_colors.append(row_colors)
            except Exception as e:
                # openpyxl 读取失败（可能是无效 XML），回退到 pandas 读取
                logger.warning(f"openpyxl 读取 .xlsx 文件失败，回退到 pandas: {e}")
                df_preview = self._read_excel_file(excel_file_path, sheet_name=sheet_name, header=None)
                
                max_check_rows = min(20, len(df_preview))
                max_cols = len(df_preview.columns) if len(df_preview) > 0 else 0
                
                rows_data = []
                rows_colors = []
                for i in range(max_check_rows):
                    row_values = [str(v) if pd.notna(v) else "" for v in df_preview.iloc[i].tolist()]
                    row_colors = [None] * len(row_values)  # 无法获取颜色，全部设为 None
                    rows_data.append(row_values)
                    rows_colors.append(row_colors)

        if self.llm_client:
            try:
                header_rows = self._detect_header_rows_with_llm_and_color(
                    rows_data, rows_colors
                )
                if header_rows:
                    color_info = {
                        "rows_data": rows_data,
                        "rows_colors": rows_colors,
                        "header_rows": header_rows,
                        "max_cols": max_cols,
                    }
                    return header_rows, color_info
            except Exception as e:
                logger.warning(f"LLM检测失败: {e}")

        df_raw = pd.DataFrame(rows_data)
        header_rows = self._detect_header_rows_rule_based(df_raw)
        color_info = {
            "rows_data": rows_data,
            "rows_colors": rows_colors,
            "header_rows": header_rows,
            "max_cols": max_cols,
        }
        return header_rows, color_info

    def _detect_header_rows_with_llm_and_color(
        self, rows_data: List[List], rows_colors: List[List]
    ) -> List[int]:
        """
        使用LLM和颜色信息检测表头行

        Args:
            rows_data: 前20行的数据
            rows_colors: 前20行的颜色信息

        Returns:
            表头行的索引列表（从0开始）
        """
        import asyncio
        import inspect

        from dbgpt.core import (
            ModelMessage,
            ModelMessageRoleType,
            ModelRequest,
            ModelRequestContext,
        )

        # 构建字典格式的数据表示（包含颜色信息）
        rows_dict = []
        for idx, (row_data, row_colors) in enumerate(
            zip(rows_data[:20], rows_colors[:20])
        ):
            # 只显示前10列数据
            row_dict = {
                "row_index": idx,
                "columns": {}
            }
            for col_idx, val in enumerate(row_data[:10]):
                col_key = f"列{col_idx + 1}"
                row_dict["columns"][col_key] = str(val) if val else ""
            
            # 统计颜色分布
            color_counts = {}
            for color in row_colors:
                if color:
                    color_counts[color] = color_counts.get(color, 0) + 1
            row_dict["color_info"] = (
                ", ".join(
                    [f"{color[:8]}({count}列)" for color, count in color_counts.items()]
                )
                if color_counts
                else "无背景色"
            )
            rows_dict.append(row_dict)
        
        rows_dict_str = json.dumps(rows_dict, ensure_ascii=False, indent=2)

        # 构建prompt
        prompt = f"""你是一个Excel数据分析专家。请分析以下Excel文件的前20行数据，判断哪些行是表头行（列名行）。

**核心判断原则**：
表头行的本质特征是：它包含**字段名/列名**（描述数据的属性），而不是数据值本身。

**判断标准（按优先级排序）**：
1. **语义判断（最重要）**：
   - 如果一行包含的是**字段名/属性名**（如"本币"、"汇率"、"姓名"、"金额"等描述性文字），而后续行包含的是**具体数据值**（如"CNY"、"1"、"张三"、"100"等），则第一行很可能是表头
   - 表头行通常：词语简短、具有描述性、表示数据的维度或属性
   - 数据行通常：包含具体数值、代码、日期等实际数据

2. **结构判断**：
   - 如果第一行看起来像列名，后续行是数据，即使没有背景色，第一行也应该是表头
   - 表头行通常在数据行之前
   - 表头行的单元格内容通常是文本性质的描述性词语

3. **可能有多级表头（多行表头）**，包括：
   - 分类标签行（如"基本信息"、"订单信息"等分类标题，通常只有少数列有内容）
   - 具体列名行（包含所有列的具体名称）
   - 多级表头时，应该包含所有相关的表头行，从分类标签行到最具体的列名行

4. **背景色（辅助判断，不是必要条件）**：
   - 表头行可能有特殊的背景色，但**没有背景色的行也可能是表头**
   - 不要因为缺少背景色就排除明显的列名行

5. **排除规则**：
   - 表头行之前可能有汇总信息行、说明行（如"请勿删除"、公式等），这些不是表头
   - 忽略包含"@@"、"@"、"="等公式标记的行，这些是Excel内部标记行

6. **中英文重复标题处理**：
   - 如果发现中英文对照的标题行（如"Name"和"中英文名"），只保留中文标题行的索引，跳过英文标题行
   - 优先选择包含中文列名的行作为表头


**重要：请按以下步骤生成结果**：
1. 首先在"reason"中**简要**说明你的判断理由（1-2句话即可），包括：
   - 哪些行是表头行，为什么
   - 如果有中英文重复标题或多级表头，简要说明处理方式
2. 然后在"header_rows"中包含应该保留的表头行索引

**注意**：reason 应该简洁明了，不要逐行分析，只需要说明最终选择的表头行及主要理由即可。

返回JSON格式：
{{
  "reason": "判断理由（必须详细说明选择的行和不选择的行，以及原因）",
  "header_rows": [行索引列表，从0开始，必须与reason中说明的选择一致]
}}

示例1（单级表头 - 无背景色但明显是列名）：
如果第0行是"本币, 本币转CNY汇率"（无背景色），第1行及之后是数据（如"CNY, 1"、"SGD, 5.4589"等，无背景色），则返回：
{{
  "reason": "第0行包含的是字段名'本币'和'本币转CNY汇率'，这些是描述性词语，表示数据的属性。第1行及之后包含的是具体数据值（如货币代码'CNY'、汇率数值'1'等）。从语义上看，第0行是列名（字段名），后续行是数据值，因此第0行是表头行。虽然都没有背景色，但基于语义判断，第0行明显是表头。",
  "header_rows": [0]
}}

示例2（多级表头）：
如果第0行是"订单信息"（有蓝色背景，只有第一列有内容），第1行是"行 ID, 订单 ID, 订单日期..."（有浅蓝色背景，所有列都有列名），第2行开始是数据（无背景色），则返回：
{{
  "reason": "第0行是分类标签'订单信息'，有蓝色背景且只有第一列有内容，属于表头的上层分类。第1行包含所有列的具体名称（行 ID, 订单 ID, 订单日期等），有浅蓝色背景，属于表头的具体列名行。这两行构成了多级表头结构，都应该保留。第2行开始无背景色且内容为数据值，不是表头。",
  "header_rows": [0, 1]
}}

示例3（表头前有说明行）：
如果第0-2行是汇总信息（无背景色或不同背景色），第3行是表头（有背景色且包含列名），则返回：
{{
  "reason": "第0-2行无背景色或背景色与表头不同，且内容为汇总信息或说明文字（不是列名），不是表头。第3行有特殊背景色且包含列名（如ID、日期等），是唯一的表头行。",
  "header_rows": [3]
}}

示例4（中英文重复标题行）：
如果第9行是"基本信息(BasicInfo)"（分类标签，有背景色），第10行是"Name, LegalName, StaffID..."（英文标题，有背景色），第11行是"@Name"（公式标记行），第12行是"中英文名, 真实姓名, 员工ID..."（中文标题，有背景色），则返回：
{{
  "reason": "第9行是分类标签'基本信息(BasicInfo)'，有背景色，属于表头的分类层。第10行是英文列名，第12行是中文列名，这两行表示相同的列。根据规则，只保留中文标题行。第11行包含'@'公式标记，应忽略。因此表头行应为第9行（分类标签）和第12行（中文列名）。",
  "header_rows": [9, 12]
}}

现在请分析以下数据（每行包含行号、列数据和颜色分布信息）：

{rows_dict_str}

请严格按照要求：先详细生成reason说明判断理由，然后基于reason生成header_rows，确保header_rows与reason完全一致。返回JSON格式的结果："""

        # 调用LLM（非流式）
        request_params = {
            "messages": [ModelMessage(role=ModelMessageRoleType.HUMAN, content=prompt)],
            "temperature": 0,
            "max_new_tokens": 1000,
            "context": ModelRequestContext(stream=False),
        }

        if self.model_name:
            request_params["model"] = self.model_name

        request = ModelRequest(**request_params)

        # 获取响应
        stream_response = self.llm_client.generate_stream(request)

        full_text = ""
        if inspect.isasyncgen(stream_response):

            async def collect_async():
                text = ""
                async for chunk in stream_response:
                    chunk_text = self._extract_chunk_text(chunk)
                    if chunk_text:
                        text = chunk_text
                return text

            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            try:
                full_text = loop.run_until_complete(collect_async())
            finally:
                loop.close()
        elif inspect.isgenerator(stream_response):
            for chunk in stream_response:
                chunk_text = self._extract_chunk_text(chunk)
                if chunk_text:
                    full_text = chunk_text
        else:
            raise Exception(f"Unexpected response type: {type(stream_response)}")

        # 解析JSON结果
        try:
            # 提取JSON部分
            start_idx = full_text.find("{")
            end_idx = full_text.rfind("}") + 1

            if start_idx >= 0 and end_idx > start_idx:
                json_str = full_text[start_idx:end_idx]
                
                # 清理控制字符（如换行符、制表符等）但不影响JSON结构
                import re
                
                # 尝试直接解析
                try:
                    result = json.loads(json_str)
                except json.JSONDecodeError as e:
                    # 如果失败，尝试清理字符串值中的控制字符
                    # JSON字符串中不允许有未转义的控制字符（U+0000到U+001F）
                    # 我们替换字符串值中的控制字符为空格
                    
                    # 方法：逐字符处理，在字符串值中替换控制字符
                    cleaned_chars = []
                    in_string = False
                    escape_next = False
                    
                    for i, char in enumerate(json_str):
                        if escape_next:
                            # 转义字符的下一个字符，直接保留
                            cleaned_chars.append(char)
                            escape_next = False
                        elif char == '\\' and in_string:
                            # 在字符串中的转义字符
                            cleaned_chars.append(char)
                            escape_next = True
                        elif char == '"' and (i == 0 or json_str[i-1] != '\\'):
                            # 字符串开始/结束（不是转义的引号）
                            in_string = not in_string
                            cleaned_chars.append(char)
                        elif in_string and ord(char) < 32 and char not in '\n\r\t':
                            # 在字符串中的未转义控制字符（除了\n\r\t），替换为空格
                            cleaned_chars.append(' ')
                        else:
                            # 其他字符直接保留
                            cleaned_chars.append(char)
                    
                    json_str_cleaned = ''.join(cleaned_chars)
                    
                    try:
                        result = json.loads(json_str_cleaned)
                    except json.JSONDecodeError:
                        # 如果还是失败，记录错误并抛出原始异常
                        logger.error(f"JSON解析失败（已尝试清理控制字符）: {e}")
                        logger.error(f"JSON字符串前500字符: {json_str[:500]}")
                        raise e

                header_rows = result.get("header_rows", [])
                reason = result.get("reason", "")

                logger.info(f"LLM判断理由: {reason}")
                logger.info(f"LLM返回的表头行: {header_rows}")

                # 验证结果
                if isinstance(header_rows, list) and all(
                    isinstance(x, int) for x in header_rows
                ):
                    # 确保索引在有效范围内
                    valid_rows = [r for r in header_rows if 0 <= r < len(rows_data)]
                    if valid_rows:
                        return sorted(valid_rows)
                    else:
                        logger.warning(f"LLM返回的表头行索引无效: {header_rows}")
                        return None
                else:
                    logger.warning(f"LLM返回的表头行格式无效: {header_rows}")
                    return None
            else:
                logger.warning(f"无法从LLM输出中提取JSON: {full_text[:200]}")
                return None
        except json.JSONDecodeError as e:
            logger.warning(f"LLM输出JSON解析失败: {e}, 输出: {full_text[:200]}")
            return None
        except Exception as e:
            logger.error(f"解析LLM结果时发生错误: {e}")
            return None

    def _merge_headers_by_color(self, color_info: Dict) -> List[str]:
        """
        基于颜色信息合并表头

        策略：
        1. 对于每个表头行，识别同一颜色的列
        2. 如果同一颜色的列中只有一个单元格有值，用该值填充其他列
        3. 将多行表头合并为单行，用"-"连接

        Args:
            color_info: 颜色信息字典

        Returns:
            合并后的表头列表
        """
        rows_data = color_info["rows_data"]
        rows_colors = color_info["rows_colors"]
        header_rows = color_info["header_rows"]
        max_cols = color_info["max_cols"]

        # 提取表头行的数据和颜色
        header_data = [rows_data[i] for i in header_rows]
        header_colors = [rows_colors[i] for i in header_rows]

        # 对每一行表头，按颜色和位置连续性分组
        filled_headers = []
        for row_idx, (row_data, row_colors) in enumerate(
            zip(header_data, header_colors)
        ):
            color_position_groups = []
            current_group = None
            current_color = None

            for col_idx, (value, color) in enumerate(zip(row_data, row_colors)):
                if color:
                    if (
                        color == current_color
                        and current_group
                        and col_idx == current_group[-1][0] + 1
                    ):
                        current_group.append((col_idx, value))
                    else:
                        if current_group:
                            color_position_groups.append((current_color, current_group))
                        current_group = [(col_idx, value)]
                        current_color = color
                else:
                    if current_group:
                        color_position_groups.append((current_color, current_group))
                        current_group = None
                        current_color = None

            if current_group:
                color_position_groups.append((current_color, current_group))

            filled_row = list(row_data)
            for color, cells in color_position_groups:
                non_empty_values = [
                    (idx, val) for idx, val in cells if val and str(val).strip()
                ]

                if len(non_empty_values) == 1:
                    fill_value = non_empty_values[0][1]
                    for col_idx, _ in cells:
                        filled_row[col_idx] = fill_value
                elif len(non_empty_values) > 1:
                    non_empty_indices = sorted([idx for idx, _ in non_empty_values])

                    for i, (val_idx, val) in enumerate(
                        sorted(non_empty_values, key=lambda x: x[0])
                    ):
                        start_idx = val_idx
                        if i < len(non_empty_indices) - 1:
                            end_idx = non_empty_indices[i + 1]
                        else:
                            end_idx = max([idx for idx, _ in cells]) + 1

                        cells_to_fill = [
                            (idx, v) for idx, v in cells if start_idx <= idx < end_idx
                        ]
                        for col_idx, _ in cells_to_fill:
                            filled_row[col_idx] = val

            filled_headers.append(filled_row)

        # 合并多级表头
        combined_headers = []
        for col_idx in range(max_cols):
            # 收集该列的所有非空值（从上层到底层）
            col_values = []
            for row_idx in range(len(filled_headers)):
                val = filled_headers[row_idx][col_idx]
                if val and str(val).strip():
                    val_str = str(val).strip()
                    # 去除换行符、普通空格和不间断空格（在合并前就清理）
                    val_str = (
                        val_str.replace("\n", "")
                        .replace("\r", "")
                        .replace("\t", "")
                        .replace(" ", "")
                        .replace("\u00a0", "")
                    )
                    # 避免重复值
                    if not col_values or val_str != col_values[-1]:
                        col_values.append(val_str)

            # 用"-"连接多级表头
            if col_values:
                combined = "-".join(col_values)
            else:
                combined = f"Column_{col_idx}"

            combined_headers.append(combined)

        return combined_headers

    def _process_multi_level_header(
        self, df_raw: pd.DataFrame, excel_file_path: str, sheet_name: str = None
    ) -> pd.DataFrame:
        """处理多级表头

        Args:
            df_raw: 原始DataFrame
            excel_file_path: Excel文件路径
            sheet_name: sheet名称
        """

        header_rows, color_info = self._detect_header_rows_with_color(
            excel_file_path, sheet_name
        )

        if not header_rows:
            header_rows = [0]

        combined_headers = self._merge_headers_by_color(color_info)

        cleaned_headers = []
        for header in combined_headers:
            cleaned = str(header)
            parts = cleaned.split("-")
            valid_parts = [
                p
                for p in parts
                if "=" not in p
                and "@@" not in p
                and "@" not in p
                and p.strip()
                and p.strip() not in ["-", "_", ""]
            ]

            if valid_parts:
                cleaned = "-".join(valid_parts)
            else:
                cleaned = f"Column_{len(cleaned_headers)}"

            cleaned = (
                cleaned.replace("\n", "")
                .replace("\r", "")
                .replace("\t", "")
                .replace(" ", "")
                .replace("\u00a0", "")
            )
            cleaned = cleaned.replace("--", "-").replace("__", "_").strip("-_")

            if not cleaned or cleaned in ["-", "_"]:
                cleaned = f"Column_{len(cleaned_headers)}"

            cleaned_headers.append(cleaned)

        final_headers = []
        seen = {}
        for header in cleaned_headers:
            if header in seen:
                seen[header] += 1
                final_headers.append(f"{header}_{seen[header]}")
            else:
                seen[header] = 0
                final_headers.append(header)

        data_start_row = max(header_rows) + 1
        data_df = df_raw.iloc[data_start_row:].copy()

        if len(final_headers) > len(data_df.columns):
            final_headers = final_headers[: len(data_df.columns)]
        elif len(final_headers) < len(data_df.columns):
            for i in range(len(final_headers), len(data_df.columns)):
                final_headers.append(f"Column_{i}")

        data_df.columns = final_headers
        data_df = data_df.dropna(how="all").reset_index(drop=True)

        return data_df

    def _clean_excel_formula(self, text: str) -> str:
        """清理Excel公式和特殊字符"""
        if not text:
            return text

        text_str = str(text)

        if text_str.startswith("="):
            import re

            quoted_texts = re.findall(r'["\']([^"\']+)["\']', text_str)
            if quoted_texts:
                text_str = "".join(quoted_texts)
            else:
                cleaned = re.sub(r"[=&()]", "", text_str)
                cleaned = re.sub(
                    r"CHAR\s*\(\s*\d+\s*\)", "", cleaned, flags=re.IGNORECASE
                )
                cleaned = re.sub(
                    r"CONCATENATE\s*\([^)]*\)", "", cleaned, flags=re.IGNORECASE
                )
                cleaned = re.sub(r"[^a-zA-Z0-9\u4e00-\u9fff_]", "", cleaned)
                if not cleaned:
                    return ""
                text_str = cleaned

        if "@@" in text_str or text_str.startswith("@"):
            import re

            text_str = re.sub(r"@@[^\u4e00-\u9fff-]*", "", text_str)
            text_str = text_str.replace("@", "")
            if not text_str.strip():
                return ""

        text_str = (
            text_str.replace("\n", "")
            .replace("\r", "")
            .replace("\t", "")
            .replace(" ", "")
        )
        return text_str

    def _remove_empty_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """删除完全为空的列"""
        df_cleaned = df.dropna(axis=1, how="all")

        empty_cols = [
            col
            for col in df_cleaned.columns
            if df_cleaned[col]
            .apply(lambda x: pd.isna(x) or (isinstance(x, str) and x.strip() == ""))
            .all()
        ]

        if empty_cols:
            df_cleaned = df_cleaned.drop(columns=empty_cols)

        return df_cleaned

    def _remove_duplicate_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """删除列名和数据值都完全重复的列"""
        columns_to_remove = []

        for i, col1 in enumerate(df.columns):
            if col1 in columns_to_remove:
                continue

            for col2 in df.columns[i + 1 :]:
                if col2 in columns_to_remove or col1 != col2:
                    continue

                try:
                    if df[col1].equals(df[col2]):
                        columns_to_remove.append(col2)
                except Exception:
                    try:
                        if (
                            df[col1]
                            .fillna("__NULL__")
                            .equals(df[col2].fillna("__NULL__"))
                        ):
                            columns_to_remove.append(col2)
                    except Exception:
                        pass

        if columns_to_remove:
            df_cleaned = df.drop(columns=columns_to_remove)
        else:
            df_cleaned = df

        return df_cleaned

    def _format_date_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """格式化日期列为YYYY-MM-DD格式"""
        df_formatted = df.copy()

        for col in df_formatted.columns:
            if pd.api.types.is_datetime64_any_dtype(df_formatted[col]):
                df_formatted[col] = df_formatted[col].apply(
                    lambda x: x.strftime("%Y-%m-%d") if pd.notna(x) else None
                )
            elif df_formatted[col].dtype == "object":
                non_null_values = df_formatted[col].dropna()
                if len(non_null_values) > 0:
                    sample_val = non_null_values.iloc[0]
                    if isinstance(sample_val, (pd.Timestamp, datetime)):
                        df_formatted[col] = df_formatted[col].apply(
                            lambda x: x.strftime("%Y-%m-%d")
                            if pd.notna(x) and isinstance(x, (pd.Timestamp, datetime))
                            else x
                        )

        return df_formatted

    def _detect_id_columns_with_llm(self, df: pd.DataFrame, table_name: str) -> List[str]:
        """使用 LLM 识别 ID 列"""
        if not self.llm_client:
            raise ValueError("LLM客户端未配置，无法识别ID列")
        
        import asyncio
        import inspect
        from dbgpt.core import ModelMessage, ModelMessageRoleType, ModelRequest
        
        # 构建列信息
        columns_info = []
        for col in df.columns:
            sample_values = df[col].dropna().head(5).tolist()
            sample_str = ", ".join([str(v) for v in sample_values])[:100]
            columns_info.append(f"  - {col} (唯一值: {len(df[col].dropna().unique())}, 示例: {sample_str})")
        
        prompt = f"""分析数据表字段，识别ID列（标识符列，如员工ID、订单号、编码等）。

表名: {table_name}
字段:
{chr(10).join(columns_info)}

返回JSON: {{"id_columns": ["列名1", "列名2"]}}
无ID列返回: {{"id_columns": []}}"""

        request = ModelRequest(
            model=self.model_name,
            messages=[ModelMessage(role=ModelMessageRoleType.HUMAN, content=prompt)],
            temperature=0,
            max_new_tokens=500,
        )
        stream_response = self.llm_client.generate_stream(request)
        
        full_text = ""
        if inspect.isasyncgen(stream_response):
            async def collect():
                text = ""
                async for chunk in stream_response:
                    text = self._extract_chunk_text(chunk) or text
                return text
            loop = asyncio.new_event_loop()
            try:
                full_text = loop.run_until_complete(collect())
            finally:
                loop.close()
        elif inspect.isgenerator(stream_response):
            for chunk in stream_response:
                full_text = self._extract_chunk_text(chunk) or full_text
        
        try:
            start, end = full_text.find("{"), full_text.rfind("}") + 1
            if start >= 0 and end > start:
                result = json.loads(full_text[start:end])
                return [col for col in result.get("id_columns", []) if col in df.columns]
        except Exception as e:
            logger.warning(f"解析ID列JSON失败: {e}")
        return []
    
    def _convert_id_columns_to_string(self, df: pd.DataFrame, id_columns: List[str]) -> pd.DataFrame:
        """将 ID 列转换为字符串类型"""
        df_converted = df.copy()
        
        def convert_to_str(x):
            if pd.isna(x):
                return None
            if isinstance(x, float) and x == int(x):
                return str(int(x))
            return str(x)
        
        for col in id_columns:
            if col in df_converted.columns:
                try:
                    df_converted[col] = df_converted[col].apply(convert_to_str).astype('object')
                except Exception as e:
                    logger.warning(f"转换ID列 '{col}' 失败: {e}")
        
        return df_converted

    def _convert_column_types(self, df: pd.DataFrame, id_columns: List[str] = None) -> pd.DataFrame:
        """
        根据字段的实际值进行智能类型转换

        转换策略：
        1. 尝试转换为日期类型
        2. 尝试转换为数值类型（整数或浮点数）
        3. 如果都失败，保持字符串类型

        Args:
            df: 原始DataFrame
            id_columns: ID 列名列表，这些列跳过数值转换

        Returns:
            转换后的DataFrame
        """
        if id_columns is None:
            id_columns = []
        
        df_converted = df.copy()

        for col in df_converted.columns:
            # 跳过 ID 列
            if col in id_columns:
                continue
            
            # 跳过已经是数值类型的列
            if df_converted[col].dtype in ["int64", "float64", "int32", "float32"]:
                continue

            # 跳过已经是日期类型的列
            if pd.api.types.is_datetime64_any_dtype(df_converted[col]):
                continue

            # 只处理object类型的列
            if df_converted[col].dtype == "object":
                non_null_data = df_converted[col].dropna()

                if len(non_null_data) == 0:
                    continue

                # 策略1: 尝试转换为日期类型
                # 检查是否有足够的数据看起来像日期（至少30%的非空值能解析为日期）
                date_success_count = 0
                try:
                    # 尝试解析前100个非空值
                    sample_size = min(100, len(non_null_data))
                    sample_data = non_null_data.head(sample_size)

                    for val in sample_data:
                        if pd.notna(val):
                            try:
                                pd.to_datetime(str(val), errors="raise")
                                date_success_count += 1
                            except Exception:
                                pass

                    # 如果超过30%的值能解析为日期，则转换整个列为日期
                    if date_success_count > sample_size * 0.3:
                        try:
                            df_converted[col] = pd.to_datetime(
                                df_converted[col], errors="coerce"
                            )
                            # 转换为字符串格式 YYYY-MM-DD
                            df_converted[col] = df_converted[col].apply(
                                lambda x: x.strftime("%Y-%m-%d")
                                if pd.notna(x)
                                else None
                            )
                            logger.debug(f"列 '{col}' 转换为日期类型")
                            continue
                        except Exception:
                            pass
                except Exception:
                    pass

                # 策略2: 尝试转换为数值类型
                # 检查是否有足够的数据看起来像数字（至少50%的非空值能解析为数字）
                numeric_success_count = 0
                has_decimal = False

                try:
                    sample_size = min(100, len(non_null_data))
                    sample_data = non_null_data.head(sample_size)

                    for val in sample_data:
                        if pd.notna(val):
                            val_str = str(val).strip()
                            # 移除常见的数字格式字符（千位分隔符、货币符号等）
                            val_str = (
                                val_str.replace(",", "")
                                .replace("￥", "")
                                .replace("$", "")
                                .replace("€", "")
                                .replace(" ", "")
                            )

                            # 检查是否为数字
                            try:
                                float_val = float(val_str)
                                numeric_success_count += 1
                                # 检查是否有小数部分
                                if "." in str(val) and float_val != int(float_val):
                                    has_decimal = True
                            except Exception:
                                pass

                    # 如果超过50%的值能解析为数字，则转换整个列为数值类型
                    if numeric_success_count > sample_size * 0.5:
                        try:
                            # 先转换为字符串，移除格式字符，再转换为数值
                            df_converted[col] = df_converted[col].astype(str)
                            df_converted[col] = (
                                df_converted[col]
                                .str.replace(",", "")
                                .str.replace("￥", "")
                                .str.replace("$", "")
                                .str.replace("€", "")
                                .str.strip()
                            )
                            df_converted[col] = pd.to_numeric(
                                df_converted[col], errors="coerce"
                            )

                            # 根据是否有小数决定使用整数还是浮点数
                            if not has_decimal and df_converted[col].notna().any():
                                # 检查所有非空值是否都是整数
                                all_integers = True
                                for val in df_converted[col].dropna():
                                    if pd.notna(val) and val != int(val):
                                        all_integers = False
                                        break

                                if all_integers:
                                    df_converted[col] = df_converted[col].astype(
                                        "Int64"
                                    )  # 可空整数类型
                                    logger.debug(f"列 '{col}' 转换为整数类型")
                                else:
                                    df_converted[col] = df_converted[col].astype(
                                        "float64"
                                    )
                                    logger.debug(f"列 '{col}' 转换为浮点数类型")
                            else:
                                df_converted[col] = df_converted[col].astype("float64")
                                logger.debug(f"列 '{col}' 转换为浮点数类型")

                            continue
                        except Exception as e:
                            logger.debug(f"列 '{col}' 转换为数值类型失败: {e}")
                            pass
                except Exception as e:
                    logger.debug(f"列 '{col}' 数值类型检测失败: {e}")
                    pass

                # 策略3: 保持为字符串类型（object）
                # 不做任何转换

        return df_converted

    def _format_numeric_columns(self, df: pd.DataFrame, id_columns: List[str] = None) -> pd.DataFrame:
        """
        格式化数值列为两位小数

        Args:
            df: 原始DataFrame
            id_columns: ID 列名列表，这些列跳过格式化

        Returns:
            格式化后的DataFrame（数值列保留两位小数）
        """
        if id_columns is None:
            id_columns = []
        
        df_formatted = df.copy()

        for col in df_formatted.columns:
            # 跳过 ID 列
            if col in id_columns:
                continue
            
            # 只处理数值类型的列
            if df_formatted[col].dtype in [
                "int64",
                "float64",
                "int32",
                "float32",
                "Int64",
            ]:
                try:
                    # 将数值列转换为浮点数，并保留两位小数
                    # 使用 apply 确保所有值都格式化为两位小数（包括整数）
                    df_formatted[col] = pd.to_numeric(
                        df_formatted[col], errors="coerce"
                    ).apply(lambda x: round(float(x), 2) if pd.notna(x) else x)
                    # 确保数据类型为 float64，这样 DuckDB 会正确存储为浮点数
                    df_formatted[col] = df_formatted[col].astype("float64")
                    logger.debug(f"列 '{col}' 格式化为两位小数 (float64)")
                except Exception as e:
                    logger.warning(f"格式化列 '{col}' 失败: {e}")
                    # 如果格式化失败，保持原样
                    pass

        return df_formatted

    def _detect_header_rows_rule_based(self, df_raw: pd.DataFrame) -> List[int]:
        """基于规则的表头行检测"""

        header_keywords = [
            "id",
            "ID",
            "编号",
            "序号",
            "行",
            "订单",
            "日期",
            "date",
            "Date",
            "时间",
            "time",
            "Time",
            "名称",
            "name",
            "Name",
            "客户",
            "产品",
            "金额",
            "价格",
            "price",
            "Price",
            "销售额",
            "利润",
            "数量",
            "quantity",
            "Quantity",
            "类别",
            "category",
            "Category",
            "类型",
            "type",
            "Type",
            "区域",
            "region",
            "Region",
            "城市",
            "city",
            "City",
            "信息",
            "info",
            "Info",
            "数据",
            "data",
            "Data",
        ]

        max_check_rows = min(20, len(df_raw))
        candidate_rows = []

        for i in range(max_check_rows):
            row = df_raw.iloc[i]
            non_null_count = row.notna().sum()
            row_text = " ".join([str(val) for val in row if pd.notna(val)])
            keyword_matches = sum(
                1 for keyword in header_keywords if keyword.lower() in row_text.lower()
            )
            score = non_null_count + keyword_matches * 2

            if score > 0:
                candidate_rows.append((i, score, non_null_count, keyword_matches))

        if not candidate_rows:
            return [0]

        candidate_rows.sort(key=lambda x: x[1], reverse=True)
        main_header_row = candidate_rows[0][0]
        header_rows = [main_header_row]

        for offset in range(1, min(4, main_header_row + 1)):
            check_row = main_header_row - offset
            if check_row >= 0:
                row = df_raw.iloc[check_row]
                if row.notna().sum() >= 2:
                    header_rows.insert(0, check_row)
                else:
                    break

        return sorted(header_rows)

    def _merge_multiple_sheets(
        self,
        sheets_data: List[Tuple[str, pd.DataFrame]],
        source_column_name: str = "数据类型",
    ) -> pd.DataFrame:
        """
        合并多个sheet的数据，添加来源标识列

        Args:
            sheets_data: [(sheet_name, dataframe), ...] 列表
            source_column_name: 来源列的列名，默认为"数据类型"

        Returns:
            合并后的DataFrame
        """
        if not sheets_data:
            raise ValueError("sheets_data不能为空")

        if len(sheets_data) == 1:
            # 只有一个sheet，直接添加来源列
            sheet_name, df = sheets_data[0]
            df_copy = df.copy()
            df_copy[source_column_name] = sheet_name
            return df_copy

        # 多个sheet的情况
        merged_dfs = []

        # 收集所有列名（按出现顺序）
        all_columns = []
        seen_columns = set()
        for sheet_name, df in sheets_data:
            for col in df.columns:
                if col not in seen_columns:
                    all_columns.append(col)
                    seen_columns.add(col)

        logger.info(f"合并{len(sheets_data)}个sheet，共{len(all_columns)}个唯一列")

        # 对每个sheet进行列对齐
        for sheet_name, df in sheets_data:
            df_copy = df.copy()

            # 添加缺失的列（填充为None）
            for col in all_columns:
                if col not in df_copy.columns:
                    df_copy[col] = None

            # 按统一的列顺序重新排列
            df_copy = df_copy[all_columns]

            # 添加来源列
            df_copy[source_column_name] = sheet_name

            merged_dfs.append(df_copy)
            logger.debug(f"Sheet '{sheet_name}': {len(df)}行 -> 对齐后{len(df_copy)}行")

        # 合并所有DataFrame
        merged_df = pd.concat(merged_dfs, ignore_index=True)
        logger.info(f"合并完成：总行数 {len(merged_df)}")

        return merged_df

    def get_table_preview_data(
        self, db_path: str, table_name: str, limit: int = None, file_name: str = None
    ) -> Dict:
        """
        从数据库中获取表格预览数据
        
        Args:
            db_path: 数据库路径
            table_name: 表名
            limit: 获取的行数限制，None表示不限制
            file_name: 文件名（可选），用于在前端显示
            
        Returns:
            包含列定义和数据行的字典
        """
        import duckdb
        
        try:
            conn = duckdb.connect(db_path, read_only=True)
            
            # 获取列定义
            columns_result = conn.execute(f'DESCRIBE "{table_name}"').fetchall()
            columns = [
                {
                    "field": col[0],
                    "type": col[1],
                    "headerName": col[0],
                }
                for col in columns_result
            ]
            
            # 获取总行数
            total_count = conn.execute(
                f'SELECT COUNT(*) FROM "{table_name}"'
            ).fetchone()[0]
            
            # 获取数据行（强制限制，防止内存溢出）
            # 即使 limit 为 None，也默认最多返回1000行用于预览
            actual_limit = min(limit if limit is not None else 1000, 1000)
            print(f"[DEBUG] get_table_preview_data: 读取预览数据，限制 {actual_limit} 行")
            
            rows_result = conn.execute(
                f'SELECT * FROM "{table_name}" LIMIT {actual_limit}'
            ).fetchall()
            
            # 转换为字典列表
            rows = []
            for idx, row in enumerate(rows_result):
                row_dict = {"id": idx}  # 为每行添加唯一ID
                for col_idx, col_info in enumerate(columns):
                    value = row[col_idx]
                    # 转换为可JSON序列化的格式
                    row_dict[col_info["field"]] = self._convert_to_json_serializable(value)
                rows.append(row_dict)
            
            conn.close()
            
            result = {
                "columns": columns,
                "rows": rows,
                "total": total_count,
            }
            
            # 如果提供了文件名，添加到结果中
            if file_name:
                result["file_name"] = file_name
            
            return result
            
        except Exception as e:
            logger.error(f"获取表格预览数据失败: {e}")
            return {
                "columns": [],
                "rows": [],
                "total": 0,
                "error": str(e),
            }

    def _generate_create_table_sql(self, db_path: str, table_name: str) -> str:
        """
        生成建表SQL语句
        
        Args:
            db_path: 数据库路径
            table_name: 表名
            
        Returns:
            建表SQL语句
        """
        import duckdb
        
        try:
            conn = duckdb.connect(db_path, read_only=True)
            columns_result = conn.execute(f'DESCRIBE "{table_name}"').fetchall()
            conn.close()
            
            columns_sql = []
            for col in columns_result:
                col_name = col[0]
                col_type = col[1]
                columns_sql.append(f'    "{col_name}" {col_type}')
            
            create_sql = f'CREATE TABLE "{table_name}" (\n' + ",\n".join(columns_sql) + "\n);"
            return create_sql
        except Exception as e:
            logger.error(f"生成建表SQL失败: {e}")
            return ""

    def process_excel_multi_tables(
        self,
        excel_file_path: str,
        force_reimport: bool = False,
        original_filename: str = None,
        conv_uid: str = None,
        sheet_names: List[str] = None,
        preview_limit: int = None,
    ) -> Dict:
        """处理Excel文件，将每个sheet存为独立的表（多表模式）

        Args:
            excel_file_path: Excel文件路径
            force_reimport: 是否强制重新导入
            original_filename: 原始文件名（可选）
            conv_uid: 会话ID（可选）
            sheet_names: 要处理的sheet名称列表，如果为None则处理所有sheet
            preview_limit: 预览数据行数限制，None表示不限制
            
        Returns:
            包含多表信息的字典，结构如下：
            {
                "status": "imported" | "cached",
                "message": "...",
                "file_hash": "文件哈希",
                "db_name": "数据库名",
                "db_path": "数据库路径",
                "tables": [
                    {
                        "sheet_name": "sheet名称",
                        "table_name": "表名",
                        "table_hash": "表哈希",
                        "row_count": 行数,
                        "column_count": 列数,
                        "columns_info": [...],
                        "data_schema_json": "...",
                        "create_table_sql": "建表SQL",
                        ...
                    },
                    ...
                ],
                "conv_uid": "会话ID"
            }
        """
        if original_filename is None:
            original_filename = Path(excel_file_path).name

        # 读取Excel获取sheet信息 - 优化：不加载整个文件
        file_ext = Path(excel_file_path).suffix.lower()
        print(f"[DEBUG] process_excel_multi_tables 开始，文件: {excel_file_path}")
        
        # 优化：使用 openpyxl 读取 sheet 名称，不加载数据
        if file_ext in ['.xlsx', '.xlsm']:
            print(f"[DEBUG] 使用 openpyxl 读取 sheet 名称...")
            from openpyxl import load_workbook
            wb = load_workbook(excel_file_path, read_only=True, data_only=True)
            all_sheet_names = wb.sheetnames
            wb.close()
            print(f"[DEBUG] 读取到 {len(all_sheet_names)} 个 sheet: {all_sheet_names}")
        elif file_ext == '.xls':
            print(f"[DEBUG] .xls 文件，使用 xlrd 读取 sheet 名称...")
            import xlrd
            book = xlrd.open_workbook(excel_file_path, on_demand=True)
            all_sheet_names = book.sheet_names()
            book.release_resources()
            print(f"[DEBUG] 读取到 {len(all_sheet_names)} 个 sheet: {all_sheet_names}")
        else:
            print(f"[DEBUG] ⚠️ 未知文件格式 {file_ext}，fallback 到 pandas")
            excel_file = pd.ExcelFile(excel_file_path)
            all_sheet_names = excel_file.sheet_names

        # 确定要处理的sheet
        if sheet_names is None:
            target_sheets = all_sheet_names
        else:
            target_sheets = []
            for name in sheet_names:
                if name in all_sheet_names:
                    target_sheets.append(name)
                else:
                    logger.warning(f"Sheet '{name}' 不存在，跳过")

            if not target_sheets:
                raise ValueError(f"指定的sheet都不存在。可用的sheet: {all_sheet_names}")

        # 计算文件级别的哈希
        file_hash = self.cache_manager.calculate_file_hash(excel_file_path)
        
        # 去除 Excel 文件的筛选状态
        excel_file_path = self._remove_excel_filters(excel_file_path)

        # 检查缓存（多表模式）
        if not force_reimport:
            cached_tables = self.cache_manager.get_tables_by_file_hash(file_hash)
            if cached_tables:
                # 检查所有缓存的表是否都存在
                first_table = cached_tables[0]
                if os.path.exists(first_table["db_path"]):
                    logger.info(f"多表缓存命中: {original_filename}, {len(cached_tables)}个表")
                    
                    # 获取每个表的预览数据
                    for table_info in cached_tables:
                        table_info["preview_data"] = self.get_table_preview_data(
                            table_info["db_path"],
                            table_info["table_name"],
                            preview_limit,
                            original_filename
                        )
                    
                    return {
                        "status": "cached",
                        "message": f"使用缓存数据，共{len(cached_tables)}个表",
                        "file_hash": file_hash,
                        "db_name": first_table["db_name"],
                        "db_path": first_table["db_path"],
                        "tables": cached_tables,
                        "conv_uid": conv_uid,
                    }

        # 没有缓存或强制重新导入
        logger.info(f"处理Excel（多表模式）: {original_filename}, {len(target_sheets)}个sheet")

        # 获取LLM客户端
        if self.llm_client is None or self.model_name is None:
            llm_client, model_name = self._get_llm_client_and_model()
            if self.llm_client is None and llm_client is not None:
                self.llm_client = llm_client
            if self.model_name is None and model_name is not None:
                self.model_name = model_name

        # 创建数据库
        db_name = f"excel_{file_hash[:8]}"
        db_filename = f"{db_name}.duckdb"
        db_path = str(self.db_storage_dir / db_filename)

        import duckdb
        
        # 删除已存在的数据库文件（如果强制重新导入）
        if force_reimport and os.path.exists(db_path):
            os.remove(db_path)
            # 同时删除缓存记录
            self.cache_manager.delete_tables_by_file_hash(file_hash)

        tables_info = []
        tables_basic_info = []  # 存储每个表的基础信息（用于统一生成schema）
        
        # 第一阶段：处理所有sheet的数据，生成基础信息
        print(f"[DEBUG] 第一阶段：处理{len(target_sheets)}个sheet的数据（使用分块模式）")
        logger.info(f"第一阶段：处理{len(target_sheets)}个sheet的数据")
        for idx, sheet_name in enumerate(target_sheets):
            logger.info(f"处理Sheet: {sheet_name}")
            print(f"[DEBUG] 开始处理 sheet: {sheet_name}")
            
            try:
                # 策略：使用完整文件读取（但会内存限制），需要区分文件大小
                file_size_mb = os.path.getsize(excel_file_path) / (1024 * 1024)
                print(f"[DEBUG] 文件大小: {file_size_mb:.2f} MB")
                
                if file_size_mb > 50:  # 大于50MB
                    print(f"[DEBUG] 大文件模式：先读取样本分析表头，再分块导入...")
                    # 第一步：读取样本用于表头分析
                    df_sample = pd.read_excel(
                        excel_file_path, 
                        sheet_name=sheet_name, 
                        header=None,
                        nrows=100
                    )
                    df = self._process_multi_level_header(df_sample, excel_file_path, sheet_name)
                    print(f"[DEBUG] 表头识别完成，列数: {len(df.columns)}")
                    
                    # 第二步：分块读取完整数据（标记：需要后续特殊处理）
                    # 为了简化，暂时跳过分块导入，使用 ExcelReader 的能力
                    print(f"[DEBUG] ⚠️ 大文件检测，建议使用 ExcelReader 分块导入")
                    # 这里先用完整读取，后面再优化
                    df_full = self._read_excel_file(excel_file_path, sheet_name=sheet_name, header=None)
                    df = self._process_multi_level_header(df_full, excel_file_path, sheet_name)
                    print(f"[DEBUG] 完整数据读取完成: {len(df)} 行")
                else:
                    print(f"[DEBUG] 小文件模式：直接读取完整数据...")
                    df_raw = self._read_excel_file(
                        excel_file_path, sheet_name=sheet_name, header=None
                    )
                    print(f"[DEBUG] 数据读取完成: {len(df_raw)} 行")
                    df = self._process_multi_level_header(df_raw, excel_file_path, sheet_name)
                    print(f"[DEBUG] 表头处理完成，列数: {len(df.columns)}")
                
                # 生成表名
                safe_sheet_name = "".join(
                    c if c.isalnum() or c == "_" else "_" for c in sheet_name
                )
                if safe_sheet_name and safe_sheet_name[0].isdigit():
                    safe_sheet_name = f"tbl_{safe_sheet_name}"
                if not safe_sheet_name or len(safe_sheet_name) < 2:
                    safe_sheet_name = f"sheet_{len(tables_info)}"
                table_name = safe_sheet_name
                
                # 计算表哈希（file_hash + sheet_name）
                table_hash = hashlib.sha256(
                    f"{file_hash}_{sheet_name}".encode("utf-8")
                ).hexdigest()
                
                # 数据清洗
                df = self._remove_empty_columns(df)
                df = self._remove_duplicate_columns(df)
                df = self._format_date_columns(df)
                
                # 识别ID列
                id_columns = []
                try:
                    id_columns = self._detect_id_columns_with_llm(df, table_name)
                    if id_columns:
                        logger.info(f"Sheet '{sheet_name}' ID列: {id_columns}")
                except Exception as e:
                    logger.warning(f"识别ID列失败: {e}")
                
                # 数据类型转换
                df = self._convert_id_columns_to_string(df, id_columns)
                df = self._convert_column_types(df, id_columns)
                df = self._format_numeric_columns(df, id_columns)
                
                # 清理列名
                df.columns = [
                    str(col)
                    .replace(" ", "")
                    .replace("\u00a0", "")
                    .replace("\n", "")
                    .replace("\r", "")
                    .replace("\t", "")
                    for col in df.columns
                ]
                
                # 去重列名
                final_columns = []
                seen_columns = {}
                for col in df.columns:
                    if col in seen_columns:
                        seen_columns[col] += 1
                        final_columns.append(f"{col}_{seen_columns[col]}")
                    else:
                        seen_columns[col] = 0
                        final_columns.append(col)
                df.columns = final_columns
                
                # 写入DuckDB
                conn = duckdb.connect(db_path)
                try:
                    table_name_quoted = f'"{table_name}"'
                    conn.execute(f"DROP TABLE IF EXISTS {table_name_quoted}")
                    conn.register("temp_df", df)
                    
                    # 构建SELECT语句
                    numeric_columns = [
                        col for col in df.columns
                        if df[col].dtype in ["int64", "float64", "int32", "float32", "Int64"]
                        and col not in id_columns
                    ]
                    
                    select_parts = []
                    for col in df.columns:
                        col_quoted = f'"{col}"'
                        if col in id_columns:
                            select_parts.append(f"CAST({col_quoted} AS VARCHAR) AS {col_quoted}")
                        elif col in numeric_columns:
                            select_parts.append(f"ROUND(CAST({col_quoted} AS DOUBLE), 2) AS {col_quoted}")
                        else:
                            select_parts.append(col_quoted)
                    
                    select_sql = ", ".join(select_parts)
                    conn.execute(f"CREATE TABLE {table_name_quoted} AS SELECT {select_sql} FROM temp_df")
                    conn.close()
                    logger.info(f"表 '{table_name}' 保存完成: {len(df)}行")
                except Exception as e:
                    conn.close()
                    logger.error(f"保存表 '{table_name}' 失败: {e}")
                    raise
                
                # 生成建表SQL
                create_table_sql = self._generate_create_table_sql(db_path, table_name)
                
                # 存储基础信息，稍后统一生成schema
                tables_basic_info.append({
                    "sheet_name": sheet_name,
                    "table_name": table_name,
                    "table_hash": table_hash,
                    "df": df,
                    "id_columns": id_columns,
                    "create_table_sql": create_table_sql,
                    "db_path": db_path,
                })
                
            except Exception as e:
                logger.error(f"处理Sheet '{sheet_name}' 失败: {e}")
                # 继续处理其他sheet
                continue

        if not tables_basic_info:
            raise ValueError("没有成功处理任何sheet")
        
        # 第二阶段：统一生成所有表的schema（table_description和推荐问题）
        logger.info(f"第二阶段：统一生成{len(tables_basic_info)}个表的schema")
        try:
            all_schemas = self._generate_multi_table_schemas_with_llm(
                tables_basic_info, original_filename
            )
        except Exception as e:
            logger.warning(f"LLM统一生成schema失败，使用基础schema: {e}")
            # 回退到单独生成
            all_schemas = {}
            for table_info in tables_basic_info:
                table_name = table_info["table_name"]
                df = table_info["df"]
                id_columns = table_info["id_columns"]
                all_schemas[table_name] = self._generate_basic_schema_json(df, table_name, id_columns)
        
        # 第三阶段：保存所有表的信息到缓存
        logger.info(f"第三阶段：保存{len(tables_basic_info)}个表的信息到缓存")
        for table_info in tables_basic_info:
            sheet_name = table_info["sheet_name"]
            table_name = table_info["table_name"]
            table_hash = table_info["table_hash"]
            df = table_info["df"]
            id_columns = table_info["id_columns"]
            create_table_sql = table_info["create_table_sql"]
            db_path = table_info["db_path"]
            
            # 获取该表的schema
            schema_json = all_schemas.get(table_name, self._generate_basic_schema_json(df, table_name, id_columns))
            summary_prompt = self._format_schema_as_prompt(schema_json, df, table_name)
            
            # 获取列信息
            conn = duckdb.connect(db_path, read_only=True)
            columns_result = conn.execute(f'DESCRIBE "{table_name}"').fetchall()
            columns_info = [
                {"name": col[0], "type": col[1], "dtype": str(df[col[0]].dtype)}
                for col in columns_result
            ]
            conn.close()
            
            # 保存到缓存
            self.cache_manager.save_table_cache_info(
                file_hash=file_hash,
                sheet_name=sheet_name,
                table_hash=table_hash,
                original_filename=original_filename,
                table_name=table_name,
                db_name=db_name,
                db_path=db_path,
                df=df,
                summary_prompt=summary_prompt,
                data_schema_json=schema_json,
                id_columns=id_columns,
                create_table_sql=create_table_sql,
            )
            
            # 获取预览数据
            preview_data = self.get_table_preview_data(
                db_path, table_name, preview_limit, original_filename
            )
            
            tables_info.append({
                "sheet_name": sheet_name,
                "table_name": table_name,
                "table_hash": table_hash,
                "row_count": len(df),
                "column_count": len(df.columns),
                "columns_info": columns_info,
                "summary_prompt": summary_prompt,
                "data_schema_json": schema_json,
                "id_columns": id_columns,
                "create_table_sql": create_table_sql,
                "preview_data": preview_data,
            })

        self._register_to_dbgpt(db_name, db_path, ",".join([t["table_name"] for t in tables_info]))

        return {
            "status": "imported",
            "message": f"成功导入{len(tables_info)}个表",
            "file_hash": file_hash,
            "db_name": db_name,
            "db_path": db_path,
            "tables": tables_info,
            "conv_uid": conv_uid,
        }

    def process_csv_to_duckdb(
        self,
        csv_file_path: str,
        db_path: str,
        table_name: str,
        chunk_size: int = 100000
    ) -> Tuple[int, List[str]]:
        """
        分块读取CSV文件并导入到DuckDB（内存优化版本）
        
        Args:
            csv_file_path: CSV文件路径
            db_path: DuckDB数据库路径
            table_name: 表名
            chunk_size: 每次读取的行数
            
        Returns:
            (总行数, 列名列表)
        """
        import duckdb
        
        logger.info(f"开始分块导入CSV文件到DuckDB: {csv_file_path}")
        print(f"[DEBUG] 分块导入CSV，chunk_size={chunk_size}")
        
        # 检测编码和分隔符
        encodings = ['utf-8', 'gbk', 'gb2312', 'gb18030', 'latin1']
        detected_encoding = 'utf-8'
        
        for encoding in encodings:
            try:
                pd.read_csv(csv_file_path, nrows=10, encoding=encoding)
                detected_encoding = encoding
                logger.info(f"检测到CSV编码: {encoding}")
                break
            except:
                continue
        
        # 使用DuckDB直接读取CSV（最高效的方式）
        conn = duckdb.connect(db_path)
        try:
            # DuckDB可以直接高效读取CSV文件
            table_name_quoted = f'"{table_name}"'
            
            # 先删除已存在的表
            conn.execute(f"DROP TABLE IF EXISTS {table_name_quoted}")
            
            # 使用DuckDB的read_csv函数直接创建表（自动推断类型）
            logger.info(f"使用DuckDB直接读取CSV...")
            conn.execute(f"""
                CREATE TABLE {table_name_quoted} AS 
                SELECT * FROM read_csv_auto(
                    '{csv_file_path}',
                    header=true,
                    sample_size=100000
                )
            """)
            
            # 获取行数和列信息
            row_count = conn.execute(f'SELECT COUNT(*) FROM {table_name_quoted}').fetchone()[0]
            columns_result = conn.execute(f'DESCRIBE {table_name_quoted}').fetchall()
            columns = [col[0] for col in columns_result]
            
            logger.info(f"CSV导入完成: {row_count:,} 行, {len(columns)} 列")
            print(f"[DEBUG] CSV导入完成: {row_count:,} 行")
            
            return row_count, columns
            
        except Exception as e:
            logger.error(f"DuckDB直接读取失败，尝试pandas分块读取: {e}")
            
            # 回退到pandas分块读取
            conn.execute(f"DROP TABLE IF EXISTS {table_name_quoted}")
            
            total_rows = 0
            columns = None
            
            for chunk_idx, chunk in enumerate(pd.read_csv(
                csv_file_path,
                chunksize=chunk_size,
                encoding=detected_encoding,
                low_memory=False
            )):
                if chunk_idx == 0:
                    columns = chunk.columns.tolist()
                    # 创建表
                    conn.register("temp_chunk", chunk)
                    conn.execute(f"CREATE TABLE {table_name_quoted} AS SELECT * FROM temp_chunk")
                else:
                    # 追加数据
                    conn.register("temp_chunk", chunk)
                    conn.execute(f"INSERT INTO {table_name_quoted} SELECT * FROM temp_chunk")
                
                total_rows += len(chunk)
                
                if chunk_idx % 10 == 0:
                    logger.info(f"已导入 {total_rows:,} 行...")
                    print(f"[DEBUG] 已导入 {total_rows:,} 行")
            
            logger.info(f"CSV分块导入完成: {total_rows:,} 行")
            return total_rows, columns
            
        finally:
            conn.close()

    def process_excel(
        self,
        excel_file_path: str,
        table_name: str = None,
        force_reimport: bool = False,
        original_filename: str = None,
        conv_uid: str = None,
        sheet_names: List[str] = None,
        merge_sheets: bool = False,
        source_column_name: str = "数据类型",
        preview_limit: int = None,
    ) -> Dict:
        """处理Excel/CSV文件，自动注册到数据源

        Args:
            excel_file_path: Excel/CSV文件路径
            table_name: 表名（可选）
            force_reimport: 是否强制重新导入
            original_filename: 原始文件名（可选）
            conv_uid: 会话ID（可选）
            sheet_names: 要处理的sheet名称列表，如果为None则处理所有sheet（仅Excel）
            merge_sheets: 是否合并多个sheet（如果为True，将多个sheet合并为一张表，仅Excel）
            source_column_name: 合并时添加的来源列名，默认为"数据类型"（仅Excel）
            preview_limit: 预览数据行数限制，None表示不限制
        """
        if original_filename is None:
            original_filename = Path(excel_file_path).name

        # 检查是否为CSV文件
        is_csv = self._is_csv_file(excel_file_path)
        
        if is_csv:
            logger.info(f"检测到CSV文件: {original_filename}")
            print(f"[DEBUG] 处理CSV文件: {excel_file_path}")
            return self._process_csv_file(
                excel_file_path,
                table_name,
                force_reimport,
                original_filename,
                conv_uid,
                preview_limit
            )

        # 先计算原始文件的哈希（在任何修改之前），确保相同文件产生相同哈希
        # 读取Excel获取sheet信息（用于哈希计算）- 优化：不加载整个文件
        file_ext = Path(excel_file_path).suffix.lower()
        print(f"[DEBUG] process_excel 开始，文件: {excel_file_path}")
        
        # 优化：使用 openpyxl 读取 sheet 名称，不加载数据
        if file_ext in ['.xlsx', '.xlsm']:
            print(f"[DEBUG] 使用 openpyxl 读取 sheet 名称...")
            from openpyxl import load_workbook
            wb = load_workbook(excel_file_path, read_only=True, data_only=True)
            all_sheet_names = wb.sheetnames
            wb.close()
            print(f"[DEBUG] 读取到 {len(all_sheet_names)} 个 sheet: {all_sheet_names}")
        elif file_ext == '.xls':
            print(f"[DEBUG] .xls 文件，使用 xlrd 读取 sheet 名称...")
            import xlrd
            book = xlrd.open_workbook(excel_file_path, on_demand=True)
            all_sheet_names = book.sheet_names()
            book.release_resources()
            print(f"[DEBUG] 读取到 {len(all_sheet_names)} 个 sheet: {all_sheet_names}")
        else:
            print(f"[DEBUG] ⚠️ 未知文件格式 {file_ext}，fallback 到 pandas")
            excel_file = pd.ExcelFile(excel_file_path)
            all_sheet_names = excel_file.sheet_names

        # 确定要处理的sheet
        if sheet_names is None:
            target_sheets = all_sheet_names
        else:
            # 验证指定的sheet是否存在
            target_sheets = []
            for name in sheet_names:
                if name in all_sheet_names:
                    target_sheets.append(name)
                else:
                    logger.warning(f"Sheet '{name}' 不存在，跳过")

            if not target_sheets:
                raise ValueError(f"指定的sheet都不存在。可用的sheet: {all_sheet_names}")

        # 使用文件级别的哈希（在去除筛选之前计算，确保相同文件产生相同哈希）
        content_hash = self.cache_manager.calculate_file_hash(
            excel_file_path, target_sheets if merge_sheets else None
        )
        
        # 去除 Excel 文件的筛选状态
        excel_file_path = self._remove_excel_filters(excel_file_path)

        # 检查缓存
        if not force_reimport:
            cached_info = self.cache_manager.get_cached_info(content_hash)
            if cached_info and os.path.exists(cached_info["db_path"]):
                # 缓存命中
                logger.info(f"缓存命中: {original_filename}")
                cached_schema_json = cached_info.get("data_schema_json")

                # 为了返回top_10_rows，需要从数据库读取
                try:
                    import duckdb

                    conn = duckdb.connect(cached_info["db_path"], read_only=True)
                    # 获取列名
                    columns_result = conn.execute(
                        f'DESCRIBE "{cached_info["table_name"]}"'
                    ).fetchall()
                    columns = [col[0] for col in columns_result]

                    # 获取前10行数据
                    rows = conn.execute(
                        f'SELECT * FROM "{cached_info["table_name"]}" LIMIT 10'
                    ).fetchall()
                    conn.close()

                    # 转换为字典列表
                    top_10_rows = [dict(zip(columns, row)) for row in rows]
                    top_10_rows = self._convert_to_json_serializable(top_10_rows)
                except Exception as e:
                    logger.warning(f"从数据库读取top_10_rows失败: {e}，将返回空列表")
                    top_10_rows = []

                # 获取预览数据
                preview_data = self.get_table_preview_data(
                    cached_info["db_path"],
                    cached_info["table_name"],
                    preview_limit,
                    original_filename
                )

                return {
                    "status": "cached",
                    "message": "使用缓存数据",
                    "content_hash": content_hash,
                    "db_name": cached_info["db_name"],
                    "db_path": cached_info["db_path"],
                    "table_name": cached_info["table_name"],
                    "row_count": cached_info["row_count"],
                    "column_count": cached_info["column_count"],
                    "columns_info": cached_info["columns_info"],
                    "summary_prompt": cached_info["summary_prompt"],
                    "data_schema_json": cached_schema_json,
                    "id_columns": cached_info.get("id_columns", []),
                    "top_10_rows": top_10_rows,
                    "preview_data": preview_data,
                    "access_count": cached_info["access_count"],
                    "last_accessed": cached_info["last_accessed"],
                    "conv_uid": conv_uid,
                }

        # 没有缓存或强制重新导入，需要完整处理
        logger.info(f"处理Excel: {original_filename}")

        # 只有当 llm_client 或 model_name 为空时，才尝试获取
        if self.llm_client is None or self.model_name is None:
            llm_client, model_name = self._get_llm_client_and_model()
            if self.llm_client is None and llm_client is not None:
                self.llm_client = llm_client
            if self.model_name is None and model_name is not None:
                self.model_name = model_name

        # 优化方案：改用分块读取方式，避免一次性加载整个文件到内存
        # 但由于后续代码依赖 df，我们使用更轻量的方式：
        # 1. 使用 DuckDB 直接读取（内存友好）
        # 2. 只在需要时加载样本数据到 pandas
        
        print(f"[DEBUG] 开始处理Excel文件（使用内存优化模式）...")
        
        # 生成表名
        if table_name is None:
            base_name = Path(original_filename).stem
            base_name = "".join(
                c if c.isalnum() or c == "_" else "_" for c in base_name
            )
            if base_name and base_name[0].isdigit():
                base_name = f"tbl_{base_name}"
            if not base_name or len(base_name) < 2:
                base_name = f"excel_table_{content_hash[:8]}"
            table_name = base_name

        db_name = f"excel_{content_hash[:8]}"
        db_filename = f"{db_name}.duckdb"
        db_path = str(self.db_storage_dir / db_filename)
        
        print(f"[DEBUG] 数据库路径: {db_path}, 表名: {table_name}")
        
        # 策略：使用 ExcelReader 的分块读取导入到 DuckDB
        # 然后只加载样本数据到 pandas 用于后续分析
        from dbgpt_app.scene.chat_data.chat_excel.excel_reader import ExcelReader
        
        print(f"[DEBUG] 使用 ExcelReader 分块导入数据...")
        excel_reader = ExcelReader(
            conv_uid="temp",
            file_path=excel_file_path,
            file_name=original_filename,
            read_type="direct",  # DuckDB 直接读取（已优化为分块）
            database_name=db_path,
            table_name=table_name,
            use_existing_db=False,
        )
        
        print(f"[DEBUG] 数据导入完成，开始后续处理...")
        
        # 从 DuckDB 加载**样本数据**用于列分析和 LLM 处理
        import duckdb
        conn = duckdb.connect(db_path)
        
        # 只加载前1000行样本
        df = conn.execute(f'SELECT * FROM "{table_name}" LIMIT 1000').df()
        print(f"[DEBUG] 加载样本数据: {len(df)} 行 用于列分析")
        
        # 对样本数据进行处理
        df = self._remove_empty_columns(df)
        df = self._remove_duplicate_columns(df)
        df = self._format_date_columns(df)
        
        # 使用 LLM 识别 ID 列
        id_columns = self._detect_id_columns_with_llm(df, table_name)
        if id_columns:
            logger.info(f"ID列: {id_columns}")
            print(f"[DEBUG] 检测到ID列: {id_columns}")
        
        # 对样本数据进行转换（仅用于分析）
        df = self._convert_id_columns_to_string(df, id_columns)
        df = self._convert_column_types(df, id_columns)
        df = self._format_numeric_columns(df, id_columns)

        df.columns = [
            str(col)
            .replace(" ", "")
            .replace("\u00a0", "")
            .replace("\n", "")
            .replace("\r", "")
            .replace("\t", "")
            for col in df.columns
        ]

        # 清理后再次去重列名
        final_columns = []
        seen_columns = {}
        for col in df.columns:
            if col in seen_columns:
                seen_columns[col] += 1
                final_columns.append(f"{col}_{seen_columns[col]}")
            else:
                seen_columns[col] = 0
                final_columns.append(col)
        df.columns = final_columns
        
        print(f"[DEBUG] 样本数据处理完成，列数: {len(df.columns)}")
        
        # 注意：df 现在只包含样本数据，但 DuckDB 中有完整数据
        # 后续需要从 DuckDB 获取实际行数
        
        # 数据已经在 DuckDB 中了（由 ExcelReader 导入）
        # 这里只需要获取实际行数
        import duckdb
        conn = duckdb.connect(db_path)
        try:
            actual_row_count = conn.execute(f'SELECT COUNT(*) FROM "{table_name}"').fetchone()[0]
            print(f"[DEBUG] 实际数据行数: {actual_row_count:,}")
        except Exception as e:
            logger.error(f"获取行数失败: {e}")
            actual_row_count = len(df)  # fallback到样本行数
        finally:
            conn.close()
        
        logger.info(f"数据处理完成: {table_name} ({actual_row_count:,}行)")

        # 获取列信息
        import duckdb
        conn = duckdb.connect(db_path)
        try:
            columns_result = conn.execute(f'DESCRIBE "{table_name}"').fetchall()
            columns_info = [
                {"name": col[0], "type": col[1], "dtype": col[1]}  # 使用 DuckDB 类型
                for col in columns_result
            ]
            # 获取实际行数
            actual_row_count = conn.execute(f'SELECT COUNT(*) FROM "{table_name}"').fetchone()[0]
        finally:
            conn.close()

        # LLM 调用可能失败，但不应该阻止数据导入
        try:
            schema_understanding_json = self._generate_schema_understanding_with_llm(
                df, table_name, id_columns
            )
        except Exception as llm_e:
            logger.warning(f"LLM生成schema失败，使用基础schema: {llm_e}")
            # 生成基础的 schema JSON（不包含 LLM 生成的业务理解）
            schema_understanding_json = self._generate_basic_schema_json(df, table_name, id_columns)
        
        summary_prompt = self._format_schema_as_prompt(
            schema_understanding_json, df, table_name
        )

        self.cache_manager.save_cache_info(
            content_hash=content_hash,
            original_filename=original_filename,
            table_name=table_name,
            db_name=db_name,
            db_path=db_path,
            df=df,  # 这里传入的是样本数据，但用于生成 schema
            summary_prompt=summary_prompt,
            data_schema_json=schema_understanding_json,
            id_columns=id_columns,
        )

        self._register_to_dbgpt(db_name, db_path, table_name)

        # 使用样本数据的前10行
        top_10_rows_raw = df.head(10).values.tolist()
        top_10_rows = self._convert_to_json_serializable(top_10_rows_raw)

        # 获取预览数据
        preview_data = self.get_table_preview_data(db_path, table_name, preview_limit, original_filename)

        return {
            "status": "imported",
            "message": "成功导入新数据",
            "content_hash": content_hash,
            "db_name": db_name,
            "db_path": db_path,
            "table_name": table_name,
            "row_count": actual_row_count,  # 使用实际行数
            "column_count": len(df.columns),
            "columns_info": columns_info,
            "summary_prompt": summary_prompt,
            "data_schema_json": schema_understanding_json,
            "id_columns": id_columns,
            "top_10_rows": top_10_rows,
            "preview_data": preview_data,
            "conv_uid": conv_uid,
        }

    def _generate_basic_schema_json(
        self, df: pd.DataFrame, table_name: str, id_columns: List[str] = None
    ) -> str:
        """生成基础的Schema JSON（不使用LLM，仅基于代码分析）"""
        import json
        
        if id_columns is None:
            id_columns = []
        
        columns = []
        for col in df.columns:
            dtype = str(df[col].dtype)
            null_count = df[col].isnull().sum()
            null_pct = (null_count / len(df)) * 100 if len(df) > 0 else 0
            
            # 判断是否为ID列
            is_id_column = col in id_columns
            
            # 确定数据类型
            if dtype in ["int64", "int32", "Int64"]:
                data_type = "整数"
            elif dtype in ["float64", "float32"]:
                data_type = "小数"
            elif "datetime" in dtype:
                data_type = "日期时间"
            else:
                data_type = "文本"
            
            col_info = {
                "column_name": col,
                "data_type": data_type,
                "description": f"字段 {col}",
                "null_percentage": round(null_pct, 1),
                "is_id_column": is_id_column,
            }
            
            # 添加唯一值（对于分类列）
            if dtype == "object":
                unique_vals = df[col].dropna().unique()
                if len(unique_vals) <= 20:
                    col_info["unique_values_top20"] = [str(v) for v in unique_vals[:20]]
            
            # 添加数值统计（对于数值列，但排除ID列）
            if dtype in ["int64", "int32", "Int64", "float64", "float32"] and not is_id_column:
                col_data = df[col].dropna()
                if len(col_data) > 0:
                    col_info["statistics_summary"] = (
                        f"范围: {col_data.min():.2f} ~ {col_data.max():.2f}, "
                        f"平均: {col_data.mean():.2f}"
                    )
            
            columns.append(col_info)
        
        schema = {
            "table_name": table_name,
            "table_description": f"数据表 {table_name}",
            "id_columns": id_columns,
            "row_count": len(df),
            "column_count": len(df.columns),
            "columns": columns,
        }
        
        return json.dumps(schema, ensure_ascii=False, indent=2)

    def _generate_multi_table_schemas_with_llm(
        self, tables_basic_info: List[Dict], filename: str
    ) -> Dict[str, str]:
        """统一生成多个表的Schema理解JSON（一次LLM调用生成所有表的描述和推荐问题）
        
        Args:
            tables_basic_info: 所有表的基础信息列表，每个元素包含：
                - table_name: 表名
                - df: DataFrame对象
                - id_columns: ID列列表
            filename: 文件名
            
        Returns:
            字典，key为table_name，value为schema JSON字符串
        """
        # 构建所有表的基础信息
        tables_info_for_prompt = []
        for table_info in tables_basic_info:
            table_name = table_info["table_name"]
            df = table_info["df"]
            
            er_info = self._prepare_er_info(df, table_name)
            numeric_stats = self._prepare_numeric_stats(df)
            categorical_distribution = self._prepare_categorical_distribution(df)
            sample_data = df.head(3).to_dict("records")
            
            tables_info_for_prompt.append({
                "table_name": table_name,
                "er_info": er_info,
                "numeric_stats": numeric_stats,
                "categorical_distribution": categorical_distribution,
                "sample_data": sample_data,
            })
        
        # 构建统一的prompt
        prompt = self._build_multi_table_schema_prompt(tables_info_for_prompt, filename)
        
        # 调用LLM生成所有表的schema
        llm_result = self._call_llm_for_schema(prompt)
        
        # 解析LLM返回的结果
        try:
            all_schemas_simplified = json.loads(llm_result)
        except json.JSONDecodeError as e:
            logger.error(f"解析多表schema JSON失败: {e}")
            raise
        
        # 提取共享的推荐问题
        shared_questions = {
            "suggested_questions_zh": all_schemas_simplified.get("suggested_questions_zh", []),
            "suggested_questions_en": all_schemas_simplified.get("suggested_questions_en", []),
        }
        
        # 为每个表补充技术性字段
        result = {}
        for table_info in tables_basic_info:
            table_name = table_info["table_name"]
            df = table_info["df"]
            id_columns = table_info["id_columns"]
            
            # 获取该表的简化schema
            simplified_schema = all_schemas_simplified.get(table_name, {})
            simplified_json = json.dumps(simplified_schema, ensure_ascii=False)
            
            # 补充技术性字段（传入共享的推荐问题）
            enriched_json = self._enrich_schema_json(
                simplified_json, df, table_name, id_columns, shared_questions=shared_questions
            )
            result[table_name] = enriched_json
        
        return result

    def _generate_schema_understanding_with_llm(
        self, df: pd.DataFrame, table_name: str, id_columns: List[str] = None, other_tables_info: List[Dict] = None
    ) -> str:
        """使用LLM生成Schema理解JSON（单表模式）
        
        Args:
            df: DataFrame对象
            table_name: 表名
            id_columns: ID列列表
            other_tables_info: 其他表的信息列表（多表模式下使用），每个元素包含：
                - table_name: 表名
                - table_description: 表描述（如果有）
        """
        if id_columns is None:
            id_columns = []
        if other_tables_info is None:
            other_tables_info = []
        
        er_info = self._prepare_er_info(df, table_name)
        numeric_stats = self._prepare_numeric_stats(df)
        categorical_distribution = self._prepare_categorical_distribution(df)

        prompt = self._build_schema_understanding_prompt(
            table_name=table_name,
            er_info=er_info,
            numeric_stats=numeric_stats,
            categorical_distribution=categorical_distribution,
            sample_data=df.head(3).to_dict("records"),
            other_tables_info=other_tables_info,
        )
        # 调用LLM生成简化的Schema JSON（只包含业务理解字段）
        simplified_json = self._call_llm_for_schema(prompt)

        # 通过代码补充技术性字段，生成完整的Schema JSON（传入已识别的ID列）
        enriched_json = self._enrich_schema_json(simplified_json, df, table_name, id_columns)

        return enriched_json

    def _prepare_er_info(self, df: pd.DataFrame, table_name: str) -> str:
        """准备ER信息（表结构）"""
        er_lines = [f"表名: {table_name}"]
        er_lines.append(f"行数: {len(df)}")
        er_lines.append(f"列数: {len(df.columns)}")
        er_lines.append("\n字段列表:")

        for col in df.columns:
            dtype = str(df[col].dtype)
            null_count = df[col].isnull().sum()
            null_pct = (null_count / len(df)) * 100
            er_lines.append(f"  - {col} ({dtype}, 缺失率: {null_pct:.1f}%)")

        return "\n".join(er_lines)

    def _prepare_numeric_stats(self, df: pd.DataFrame) -> str:
        """准备数值列的描述统计"""
        numeric_cols = df.select_dtypes(include=["number"]).columns.tolist()

        if not numeric_cols:
            return "无数值列"

        stats_lines = ["数值列描述统计:"]
        for col in numeric_cols:
            col_data = df[col].dropna()
            if len(col_data) > 0:
                stats_lines.append(f"\n  {col}:")
                stats_lines.append(f"    最小值: {col_data.min():.2f}")
                stats_lines.append(f"    最大值: {col_data.max():.2f}")
                stats_lines.append(f"    平均值: {col_data.mean():.2f}")
                stats_lines.append(f"    中位数: {col_data.median():.2f}")
                stats_lines.append(f"    标准差: {col_data.std():.2f}")

        return "\n".join(stats_lines)

    def _prepare_categorical_distribution(self, df: pd.DataFrame) -> str:
        """准备分类列的唯一值分布"""
        categorical_cols = df.select_dtypes(
            include=["object", "category"]
        ).columns.tolist()

        if not categorical_cols:
            return "无分类列"

        dist_lines = ["分类列唯一值分布:"]
        for col in categorical_cols:
            unique_vals = df[col].dropna().unique()
            unique_count = len(unique_vals)

            dist_lines.append(f"\n  {col} (唯一值数量: {unique_count}):")

            if unique_count <= 20:
                # 显示所有唯一值
                value_counts = df[col].value_counts()
                for val, count in value_counts.head(20).items():
                    dist_lines.append(
                        f"    - '{val}': {count}条 ({count / len(df) * 100:.1f}%)"
                    )
            else:
                # 只显示前10个最常见的值
                value_counts = df[col].value_counts()
                dist_lines.append("    前10个最常见值:")
                for val, count in value_counts.head(10).items():
                    dist_lines.append(
                        f"    - '{val}': {count}条 ({count / len(df) * 100:.1f}%)"
                    )

        return "\n".join(dist_lines)

    def _build_multi_table_schema_prompt(
        self, tables_info: List[Dict], filename: str
    ) -> str:
        """构建多表统一生成schema的prompt
        
        Args:
            tables_info: 所有表的信息列表
            filename: 文件名
        """
        # 转换sample_data中的特殊类型为可JSON序列化的格式
        def convert_to_serializable(obj):
            if isinstance(obj, dict):
                return {k: convert_to_serializable(v) for k, v in obj.items()}
            elif isinstance(obj, list):
                return [convert_to_serializable(item) for item in obj]
            elif pd.isna(obj):
                return None
            elif hasattr(obj, "isoformat"):
                return obj.isoformat()
            elif isinstance(obj, (int, float, str, bool, type(None))):
                return obj
            else:
                return str(obj)
        
        # 构建每个表的信息文本
        tables_text = []
        for idx, table_info in enumerate(tables_info, 1):
            table_name = table_info["table_name"]
            er_info = table_info["er_info"]
            numeric_stats = table_info["numeric_stats"]
            categorical_distribution = table_info["categorical_distribution"]
            sample_data = convert_to_serializable(table_info["sample_data"])
            sample_data_str = json.dumps(sample_data, ensure_ascii=False, indent=2)
            
            table_text = f"""
=== 表{idx}: {table_name} ===

{er_info}

数值列描述统计:
{numeric_stats}

分类列唯一值分布:
{categorical_distribution}

样本数据（前3行）:
{sample_data_str}
"""
            tables_text.append(table_text)
        
        all_tables_text = "\n".join(tables_text)
        all_table_names = [t["table_name"] for t in tables_info]
        all_table_names_str = "、".join(all_table_names)
        
        prompt = f"""你是一个数据分析专家。请分析Excel文件"{filename}"中的{len(tables_info)}个数据表，为每个表生成Schema理解的JSON。

{all_tables_text}

**任务要求**：
1. 为每个表生成 `table_description`（表的整体描述，说明这是什么数据，适合做什么分析）
2. 为整个Excel文件生成 `suggested_questions_zh` 和 `suggested_questions_en`（9个推荐问题）

**推荐问题要求**：
- 前6个问题：简单的问题，有明确的标准答案
- 后3个问题：中等难度问题，需要一定的思考和分析
- **重要**：所有问题必须基于数据表中的实际字段和数据，不能凭空捏造
- **多表场景特别要求**：必须在问题中明确指出要使用哪个表或哪些表，避免"有哪些人"、"谁的奖金最高"这种模糊表达
  * 好的示例："{all_table_names[0]}中有哪些人"、"所有表中谁的奖金最高"、"{all_table_names[0]}和{all_table_names[1] if len(all_table_names) > 1 else all_table_names[0]}的数据对比"
  * 不好的示例："有哪些人"、"谁的奖金最高"
- 推荐问题应该覆盖不同的表，引导用户探索各个表的数据

请严格按照以下JSON格式输出（注意：推荐问题是整个Excel文件共享的，不是每个表单独的）：

```json
{{
  "{all_table_names[0]}": {{
    "table_description": "表的整体描述...",
  }},
  "{all_table_names[1] if len(all_table_names) > 1 else 'table2'}": {{
    "table_description": "表的整体描述...",
  }},
  ...
  "suggested_questions_zh": [
    "问题1（简单问题，明确指出使用哪个表）",
    "问题2（简单问题，明确指出使用哪个表）",
    "问题3（简单问题，明确指出使用哪个表）",
    "问题4（简单问题，明确指出使用哪个表）",
    "问题5（简单问题，明确指出使用哪个表）",
    "问题6（简单问题，明确指出使用哪个表）",
    "问题7（中等难度问题）",
    "问题8（中等难度问题）",
    "问题9（中等难度问题）"
  ],
  "suggested_questions_en": [
    "Question 1 (simple question, specify which table to use)",
    "Question 2 (simple question, specify which table to use)",
    "Question 3 (simple question, specify which table to use)",
    "Question 4 (simple question, specify which table to use)",
    "Question 5 (simple question, specify which table to use)",
    "Question 6 (simple question, specify which table to use)",
    "Question 7 (medium difficulty question)",
    "Question 8 (medium difficulty question)",
    "Question 9 (medium difficulty question)"
  ]
}}
```

请直接输出JSON，不要有其他文字："""
        
        return prompt

    def _build_schema_understanding_prompt(
        self,
        table_name: str,
        er_info: str,
        numeric_stats: str,
        categorical_distribution: str,
        sample_data: list,
        other_tables_info: List[Dict] = None,
    ) -> str:
        """构建Schema理解Prompt（简化版，只生成必要的业务理解字段）

        Args:
            table_name: 当前表名
            er_info: ER信息
            numeric_stats: 数值统计
            categorical_distribution: 分类分布
            sample_data: 样本数据
            other_tables_info: 其他表的信息列表（多表模式下使用），每个元素包含：
                - table_name: 表名
                - table_description: 表描述（如果有）
        """

        # 转换sample_data中的特殊类型（如Timestamp）为可JSON序列化的格式
        def convert_to_serializable(obj):
            """递归转换对象为可JSON序列化的格式"""
            if isinstance(obj, dict):
                return {k: convert_to_serializable(v) for k, v in obj.items()}
            elif isinstance(obj, list):
                return [convert_to_serializable(item) for item in obj]
            elif pd.isna(obj):
                return None
            elif hasattr(obj, "isoformat"):  # datetime, date, time, Timestamp
                return obj.isoformat()
            elif isinstance(obj, (int, float, str, bool, type(None))):
                return obj
            else:
                return str(obj)

        serializable_sample_data = convert_to_serializable(sample_data)
        sample_data_str = json.dumps(
            serializable_sample_data, ensure_ascii=False, indent=2
        )

        # 构建多表场景的提示信息
        multi_table_guidance = ""
        if other_tables_info and len(other_tables_info) > 0:
            other_tables_names = [t.get("table_name", "") for t in other_tables_info]
            other_tables_str = "、".join(other_tables_names)
            other_tables_str_en = ", ".join(other_tables_names)
            example_other_table = other_tables_names[0] if other_tables_names else "其他表"
            multi_table_guidance = f"""

**重要提示 - 多表场景 / IMPORTANT - Multi-Table Scenario**：
当前Excel文件中包含多个数据表，除了当前表"{table_name}"外，还有以下表：{other_tables_str}。
This Excel file contains multiple data tables. Besides the current table "{table_name}", there are also the following tables: {other_tables_str_en}.

**推荐问题的特殊要求 / Special Requirements for Recommended Questions**：
1. **必须在问题中明确指出要使用哪个表或哪些表，避免模糊表达**
   **Questions must explicitly specify which table(s) to use, avoid vague expressions**
2. **推荐问题的示例格式 / Example Question Formats**：
   - ✅ 好的示例（明确指出表）/ Good Examples (explicitly specify tables)：
     * 中文: "查询{table_name}中有哪些人"、"查询{example_other_table}中有哪些人"、"所有员工中有哪些人"、"{table_name}中谁的奖金最高"
     * English: "Who are in {table_name}", "Who are in {example_other_table}", "Who are all employees", "Who has the highest bonus in {table_name}"
   - ❌ 不好的示例（没有明确指出表）/ Bad Examples (no table specified)：
     * 中文: "有哪些人"、"谁的奖金最高"、"数据有多少条"
     * English: "Who are there", "Who has the highest bonus", "How many records"

3. **推荐问题的分配建议 / Question Distribution Suggestions**：
   - 可以包含几个只针对当前表"{table_name}"的问题（明确提到表名）
   - Can include a few questions specific to the current table "{table_name}" (explicitly mention table name)
   - 可以包含几个针对其他表的问题（明确提到其他表名）
   - Can include a few questions for other tables (explicitly mention other table names)
   - 可以包含几个涉及"所有"或"全部"表的问题（使用"所有"、"全部"、"总共"等词 / 英文使用"all", "total", "overall"等词）
   - Can include a few questions involving "all" or "total" tables (use words like "所有", "全部", "总共" in Chinese / "all", "total", "overall" in English)
"""

        prompt = f"""你是一个数据分析专家，请分析以下数据表的结构和语义，生成Schema理解的JSON。

=== 数据表ER信息 ===
{er_info}

=== 数值列描述统计 ===
{numeric_stats}

=== 分类列唯一值分布 ===
{categorical_distribution}
{multi_table_guidance}

请生成一个简化的JSON格式，包含以下信息：

1. **table_description**: 表的整体描述，说明这是什么数据，适合做什么分析
2. **suggested_questions**: 生成9个推荐问题，帮助用户快速了解数据
   - **suggested_questions_zh**: 中文版本的9个推荐问题
   - **suggested_questions_en**: 英文版本的9个推荐问题（与中文问题对应，内容相同但语言不同）
   - **问题要求**：
     * 前6个问题：简单的问题，有明确的标准答案
     * 后3个问题：中等难度问题，需要一定的思考和分析
     * **重要**：所有问题必须基于数据表中的实际字段和数据，不能凭空捏造不存在的字段或数据
     * **多表场景特别要求**：必须在问题中明确指出要使用哪个表或哪些表，避免"有哪些人"、"谁的奖金最高"这种模糊表达，应该写成"{{表名}}中有哪些人"、"{{表名}}中谁的奖金最高"或"所有员工中有哪些人"等明确形式

请严格按照以下JSON格式输出：

```json
{{
  "table_description": "表的整体描述...",
  "suggested_questions_zh": [
    "问题1（简单问题，有标准答案）",
    "问题2（简单问题，有标准答案）",
    "问题3（简单问题，有标准答案）",
    "问题4（简单问题，有标准答案）",
    "问题5（简单问题，有标准答案）",
    "问题6（简单问题，有标准答案）",
    "问题7（中等难度问题）",
    "问题8（中等难度问题）",
    "问题9（中等难度问题）"
  ],
  "suggested_questions_en": [
    "Question 1 (simple question with standard answer)",
    "Question 2 (simple question with standard answer)",
    "Question 3 (simple question with standard answer)",
    "Question 4 (simple question with standard answer)",
    "Question 5 (simple question with standard answer)",
    "Question 6 (simple question with standard answer)",
    "Question 7 (medium difficulty question)",
    "Question 8 (medium difficulty question)",
    "Question 9 (medium difficulty question)"
  ]
}}
```

注意：
1. **前6个问题必须是简单的问题，有明确的标准答案**
4. **后3个问题必须是开放式问题**，可以引发深入思考和分析
5. **所有问题必须基于数据表中的实际字段和数据，可以围绕具体的分类值进行分析，不能凭空捏造不存在的字段或数据**
6. 中文推荐问题应该用自然的中文表达，英文推荐问题应该用自然的英文表达，简洁明了，可以直接用于数据分析
7. 中英文问题应该一一对应，内容相同但语言不同
8. **如果是多表场景，必须在问题中明确指出要使用哪个表或哪些表**，使用表名或"所有"、"全部"等明确表达，避免模糊的问题

请直接输出JSON，不要有其他文字：
"""
        return prompt

    def _extract_chunk_text(self, chunk) -> str:
        """统一的chunk文本提取方法"""
        try:
            if hasattr(chunk, "text"):
                return chunk.text
            elif hasattr(chunk, "content"):
                if hasattr(chunk.content, "get_text"):
                    try:
                        return chunk.content.get_text()
                    except Exception:
                        pass
                elif isinstance(chunk.content, str):
                    return chunk.content
        except Exception as e:
            logger.debug(f"提取chunk文本失败: {e}")
        return ""

    def _enrich_schema_json(
        self, simplified_json: str, df: pd.DataFrame, table_name: str, 
        pre_detected_id_columns: List[str] = None, shared_questions: Dict = None
    ) -> str:
        """通过代码补充技术性字段，生成完整的Schema JSON
        
        Args:
            simplified_json: LLM生成的简化schema JSON
            df: DataFrame对象
            table_name: 表名
            pre_detected_id_columns: 预先检测的ID列
            shared_questions: 共享的推荐问题（多表模式下使用），包含：
                - suggested_questions_zh: 中文推荐问题列表
                - suggested_questions_en: 英文推荐问题列表
        """
        if pre_detected_id_columns is None:
            pre_detected_id_columns = []
        
        try:
            schema = json.loads(simplified_json)
        except json.JSONDecodeError as e:
            logger.error(f"解析Schema JSON失败: {e}")
            raise

        id_columns = pre_detected_id_columns or []

        # 构建完整的columns列表
        enriched_columns = []
        for col_name in df.columns:
            col_data = df[col_name]
            dtype = str(col_data.dtype)

            # 判断是否为ID列
            is_id_column = col_name in id_columns
            
            col_info = {
                "column_name": col_name,
                "data_type": dtype,
                "description": self._generate_column_description(col_name, col_data, dtype),
                "is_key_field": self._is_potential_key_field(col_name, col_data),
                "is_id_column": is_id_column,
            }

            # 判断字段类型（根据数据类型判断）
            is_numeric_by_dtype = dtype in ["int64", "float64", "int32", "float32", "Int64"]

            # 数值字段：根据数据类型判断
            if is_numeric_by_dtype:
                numeric_data = col_data.dropna()

                if len(numeric_data) > 0:
                    try:
                        min_val = numeric_data.min()
                        max_val = numeric_data.max()
                        mean_val = numeric_data.mean()
                        median_val = numeric_data.median()
                        col_info["statistics_summary"] = (
                            f"范围: [{min_val:.2f}, {max_val:.2f}], "
                            f"均值: {mean_val:.2f}, 中位数: {median_val:.2f}"
                        )
                    except Exception:
                        # 如果转换失败，不添加统计信息
                        pass

            # 分类字段：列出出现次数最高的20个值
            elif dtype in ["object", "category"]:
                value_counts = col_data.value_counts()
                total_unique = len(col_data.dropna().unique())

                # 只取出现次数最高的20个值
                top_20_values = value_counts.head(20)
                col_info["unique_values_top20"] = [
                    str(v) for v in top_20_values.index.tolist()
                ]

            enriched_columns.append(col_info)

        # 从LLM返回的schema或共享问题中提取推荐问题
        if shared_questions:
            # 多表模式：使用共享的推荐问题
            suggested_questions_zh = shared_questions.get("suggested_questions_zh", [])
            suggested_questions_en = shared_questions.get("suggested_questions_en", [])
        else:
            # 单表模式：从schema中提取推荐问题
            suggested_questions_zh = schema.get("suggested_questions_zh", [])
            suggested_questions_en = schema.get("suggested_questions_en", [])
            
            # 兼容旧格式：如果只有 suggested_questions，则作为中文版本
            if not suggested_questions_zh and schema.get("suggested_questions"):
                suggested_questions_zh = schema.get("suggested_questions", [])
        
        # 如果中文版本没有或数量不足，使用备用方法生成
        if not suggested_questions_zh or len(suggested_questions_zh) < 9:
            suggested_questions_zh = self._generate_fallback_questions(
                enriched_columns, df
            )
        
        # 如果英文版本没有或数量不足，尝试从中文翻译或使用备用方法
        if not suggested_questions_en or len(suggested_questions_en) < 9:
            # 如果有中文版本，可以尝试翻译（这里先使用备用方法生成英文版本）
            suggested_questions_en = self._generate_fallback_questions_en(
                enriched_columns, df
            )

        # 获取2行样本数据用于query改写
        sample_rows = self._get_sample_rows(df, n=2)

        return json.dumps(
            {
                "table_name": table_name,
                "table_description": schema.get("table_description", ""),
                "id_columns": id_columns,  # 保留ID列信息
                "columns": enriched_columns,
                "sample_rows": sample_rows,  # 添加样本数据
                "suggested_questions_zh": suggested_questions_zh[:9],  # 确保最多9个
                "suggested_questions_en": suggested_questions_en[:9],  # 确保最多9个
            },
            ensure_ascii=False,
            indent=2,
        )

    def _generate_fallback_questions(
        self, columns: List[Dict], df: pd.DataFrame
    ) -> List[str]:
        """
        当LLM生成失败时，使用规则生成基础的推荐问题

        Args:
            columns: 字段信息列表
            df: DataFrame对象

        Returns:
            推荐问题列表
        """
        questions = []
        row_count = len(df)

        # 找出数值字段和分类字段（根据data_type判断）
        numeric_cols = [
            col.get("column_name")
            for col in columns
            if col.get("data_type") in ["int64", "float64", "int32", "float32", "Int64"]
        ]
        categorical_cols = [
            col.get("column_name")
            for col in columns
            if col.get("data_type") in ["object", "category"]
        ]
        # 时间字段：根据列名和数据类型判断
        time_cols = []
        for col in columns:
            col_name = col.get("column_name", "").lower()
            if any(keyword in col_name for keyword in ["时间", "日期", "date", "time"]):
                time_cols.append(col.get("column_name"))

        # 生成基础问题：6个简单问题 + 3个开放式问题
        # 简单问题1-6：基于实际字段
        questions.append(f"数据总共有多少条记录？")
        
        if numeric_cols:
            questions.append(f"{numeric_cols[0]}的平均值是多少？")
            if len(numeric_cols) > 1:
                questions.append(f"{numeric_cols[1]}的最大值是多少？")
            else:
                questions.append(f"{numeric_cols[0]}的最大值是多少？")
            if len(numeric_cols) > 2:
                questions.append(f"{numeric_cols[2]}的最小值是多少？")
            else:
                questions.append(f"{numeric_cols[0]}的最小值是多少？")
        else:
            questions.append(f"数据的基本统计信息是什么？")
            questions.append(f"数据的主要特征是什么？")
            questions.append(f"数据的基本分布情况如何？")
        
        if categorical_cols:
            questions.append(f"{categorical_cols[0]}有多少个不同的值？")
            if len(categorical_cols) > 1:
                questions.append(f"{categorical_cols[1]}的分布情况如何？")
            else:
                questions.append(f"{categorical_cols[0]}出现次数最多的值是什么？")
        else:
            questions.append(f"数据的主要分类维度是什么？")
            questions.append(f"数据的主要分组方式是什么？")
        
        # 开放式问题7-9：基于实际字段
        if len(numeric_cols) > 0 and len(categorical_cols) > 0:
            questions.append(f"按{categorical_cols[0]}分组，分析{numeric_cols[0]}的分布情况")
            if len(categorical_cols) > 1:
                questions.append(f"不同{categorical_cols[0]}之间的{numeric_cols[0]}有何差异？")
            else:
                questions.append(f"{categorical_cols[0]}与{numeric_cols[0]}之间的关系如何？")
        elif time_cols and len(time_cols) > 0:
            questions.append(f"按{time_cols[0]}分析数据的变化趋势")
            questions.append(f"数据在{time_cols[0]}维度上的主要变化规律是什么？")
        elif len(categorical_cols) > 1:
            questions.append(f"不同{categorical_cols[0]}之间的{categorical_cols[1]}分布有何差异？")
            questions.append(f"{categorical_cols[0]}与{categorical_cols[1]}之间的关联关系如何？")
        else:
            questions.append(f"数据的主要特征和规律是什么？")
            questions.append(f"数据中哪些因素影响了主要指标的变化？")
        
        # 最后一个开放式问题
        if len(numeric_cols) > 0:
            questions.append(f"如何优化{numeric_cols[0]}这个指标？")
        else:
            questions.append(f"如何进一步分析这些数据？")

        return questions[:9]  # 确保最多9个

    def _generate_fallback_questions_en(
        self, columns: List[Dict], df: pd.DataFrame
    ) -> List[str]:
        """
        Generate fallback suggested questions in English when LLM generation fails

        Args:
            columns: Column information list
            df: DataFrame object

        Returns:
            List of suggested questions in English
        """
        questions = []
        row_count = len(df)

        # Find numeric fields and categorical fields (based on data_type)
        numeric_cols = [
            col.get("column_name")
            for col in columns
            if col.get("data_type") in ["int64", "float64", "int32", "float32", "Int64"]
        ]
        categorical_cols = [
            col.get("column_name")
            for col in columns
            if col.get("data_type") in ["object", "category"]
        ]
        # Time fields: based on column name and data type
        time_cols = []
        for col in columns:
            col_name = col.get("column_name", "").lower()
            if any(keyword in col_name for keyword in ["时间", "日期", "date", "time"]):
                time_cols.append(col.get("column_name"))

        # Generate basic questions: 6 simple questions + 3 open-ended questions
        # Simple questions 1-6: Based on actual fields
        questions.append("How many records are there in total?")
        
        if numeric_cols:
            questions.append(f"What is the average value of {numeric_cols[0]}?")
            if len(numeric_cols) > 1:
                questions.append(f"What is the maximum value of {numeric_cols[1]}?")
            else:
                questions.append(f"What is the maximum value of {numeric_cols[0]}?")
            if len(numeric_cols) > 2:
                questions.append(f"What is the minimum value of {numeric_cols[2]}?")
            else:
                questions.append(f"What is the minimum value of {numeric_cols[0]}?")
        else:
            questions.append("What are the basic statistics of the data?")
            questions.append("What are the main characteristics of the data?")
            questions.append("What is the basic distribution of the data?")
        
        if categorical_cols:
            questions.append(f"How many distinct values are in {categorical_cols[0]}?")
            if len(categorical_cols) > 1:
                questions.append(f"What is the distribution of {categorical_cols[1]}?")
            else:
                questions.append(f"What is the most frequent value in {categorical_cols[0]}?")
        else:
            questions.append("What are the main categorical dimensions in the data?")
            questions.append("What are the main grouping methods in the data?")
        
        # Open-ended questions 7-9: Based on actual fields
        if len(numeric_cols) > 0 and len(categorical_cols) > 0:
            questions.append(f"Group by {categorical_cols[0]}, analyze the distribution of {numeric_cols[0]}")
            if len(categorical_cols) > 1:
                questions.append(f"What are the differences in {numeric_cols[0]} across different {categorical_cols[0]}?")
            else:
                questions.append(f"What is the relationship between {categorical_cols[0]} and {numeric_cols[0]}?")
        elif time_cols and len(time_cols) > 0:
            questions.append(f"Analyze the trend of data changes by {time_cols[0]}")
            questions.append(f"What are the main patterns of data changes in the {time_cols[0]} dimension?")
        elif len(categorical_cols) > 1:
            questions.append(f"What are the differences in {categorical_cols[1]} distribution across different {categorical_cols[0]}?")
            questions.append(f"What is the relationship between {categorical_cols[0]} and {categorical_cols[1]}?")
        else:
            questions.append("What are the main characteristics and patterns in the data?")
            questions.append("What factors affect the changes in the main indicators?")
        
        # Last open-ended question
        if len(numeric_cols) > 0:
            questions.append(f"How to optimize the {numeric_cols[0]} indicator?")
        else:
            questions.append("How to further analyze this data?")

        return questions[:9]  # Ensure maximum 9 questions

    def _call_llm_for_schema(self, prompt: str) -> str:
        """调用LLM生成Schema JSON"""
        try:
            import asyncio
            import logging

            from dbgpt.core import ModelMessage, ModelMessageRoleType, ModelRequest

            worker_logger = logging.getLogger(
                "dbgpt.model.cluster.worker.default_worker"
            )
            original_level = worker_logger.level

            try:
                worker_logger.setLevel(logging.ERROR)

                request_params = {
                    "messages": [
                        ModelMessage(role=ModelMessageRoleType.HUMAN, content=prompt)
                    ],
                    "temperature": 0,
                    "max_new_tokens": 20480,
                }

                if self.model_name:
                    request_params["model"] = self.model_name

                request = ModelRequest(**request_params)

                if hasattr(self.llm_client, "generate_stream"):
                    import inspect

                    stream_response = self.llm_client.generate_stream(request)

                    full_text = ""
                    if inspect.isasyncgen(stream_response):

                        async def collect_async():
                            text = ""
                            async for chunk in stream_response:
                                chunk_text = self._extract_chunk_text(chunk)
                                if chunk_text:
                                    text = chunk_text
                            return text

                        loop = asyncio.new_event_loop()
                        asyncio.set_event_loop(loop)
                        try:
                            full_text = loop.run_until_complete(collect_async())
                        finally:
                            loop.close()
                    elif inspect.isgenerator(stream_response):
                        for chunk in stream_response:
                            chunk_text = self._extract_chunk_text(chunk)
                            if chunk_text:
                                full_text = chunk_text
                    else:
                        raise Exception(
                            f"Unexpected response type: {type(stream_response)}"
                        )

                    class FakeResponse:
                        def __init__(self, text):
                            self.text = text

                    response = FakeResponse(full_text)

                else:
                    raise Exception("LLM客户端没有generate_stream方法")

            finally:
                worker_logger.setLevel(original_level)

            if response and hasattr(response, "text") and response.text:
                text = response.text.strip()
                json_str = None

                if "```json" in text.lower():
                    start_idx = text.lower().find("```json")
                    if start_idx >= 0:
                        content_start = text.find("\n", start_idx) + 1
                        if content_start > 0:
                            end_idx = text.find("```", content_start)
                            if end_idx > content_start:
                                json_str = text[content_start:end_idx].strip()

                if not json_str and "```" in text:
                    start_idx = text.find("```")
                    content_start = text.find("\n", start_idx) + 1
                    if content_start > 0:
                        end_idx = text.find("```", content_start)
                        if end_idx > content_start:
                            json_str = text[content_start:end_idx].strip()

                if not json_str:
                    start = text.find("{")
                    end = text.rfind("}") + 1
                    if start >= 0 and end > start:
                        json_str = text[start:end].strip()

                if not json_str:
                    raise Exception("无法提取JSON内容")

                try:
                    parsed = json.loads(json_str)
                    return json_str
                except json.JSONDecodeError as e:
                    logger.error(f"JSON格式错误: {e}")
                    raise
            else:
                raise Exception("LLM返回空结果")

        except Exception as e:
            logger.error(f"调用LLM失败: {e}")
            raise

    def _format_schema_as_prompt(
        self, schema_json: str, df: pd.DataFrame, table_name: str
    ) -> str:
        """
        将Schema JSON格式化为文本prompt
        用于后续的query改写和SQL生成
        """
        try:
            schema = json.loads(schema_json)
        except json.JSONDecodeError:
            # 如果JSON解析失败，返回简单格式
            return f"数据表: {table_name}\n数据规模: {len(df)}行 × {len(df.columns)}列"

        # 构建易读的文本格式
        lines = []
        lines.append("=== 数据表Schema理解 ===")
        lines.append(f"表名: {schema.get('table_name', table_name)}")
        lines.append(f"表描述: {schema.get('table_description', '')}")
        lines.append("")

        lines.append("=== 字段详细信息 ===")
        for col in schema.get("columns", []):
            lines.append(f"\n字段: {col.get('column_name')}")
            lines.append(f"  类型: {col.get('data_type')}")
            lines.append(f"  描述: {col.get('description')}")

            if "unique_values_top20" in col:
                unique_vals = col["unique_values_top20"]
                lines.append(
                    f"  出现次数前20的值: {', '.join([str(v) for v in unique_vals])}"
                )

            if "statistics_summary" in col:
                lines.append(f"  统计: {col['statistics_summary']}")

        return "\n".join(lines)

    def _generate_column_description(
        self, col_name: str, col_data: pd.Series, dtype: str
    ) -> str:
        """
        根据规则生成字段描述（简化版）
        
        Args:
            col_name: 字段名
            col_data: 字段数据
            dtype: 数据类型
            
        Returns:
            字段描述（不包含字段名，避免重复）
        """
        # 根据数据类型生成基础描述
        if dtype in ["int64", "float64", "int32", "float32", "Int64"]:
            return "数值类型"
        elif "date" in dtype.lower() or "time" in dtype.lower():
            return "日期时间类型"
        elif dtype in ["object", "category"]:
            # 对于分类字段，显示唯一值数量
            unique_count = len(col_data.dropna().unique())
            if unique_count <= 20:
                return f"分类字段，{unique_count}个可选值"
            else:
                return "文本类型"
        else:
            return f"{dtype}类型"
    
    def _get_sample_values(self, col_data: pd.Series, n: int = 3) -> list:
        """
        获取列的示例值
        """
        non_null_values = col_data.dropna().unique()
        if len(non_null_values) > 0:
            sample = non_null_values[:n].tolist()
            return [str(v) for v in sample]
        return []

    def _get_sample_rows(self, df: pd.DataFrame, n: int = 2) -> list:
        """
        获取n行样本数据，用于query改写时提供真实数据参考
        
        Args:
            df: DataFrame对象
            n: 样本行数
            
        Returns:
            样本数据列表，每行是一个字典
        """
        try:
            # 取前n行数据
            sample_df = df.head(n)
            rows = []
            for _, row in sample_df.iterrows():
                row_dict = {}
                for col in sample_df.columns:
                    val = row[col]
                    # 转换为可序列化的格式
                    if pd.isna(val):
                        row_dict[col] = None
                    elif isinstance(val, (pd.Timestamp, datetime)):
                        row_dict[col] = val.strftime("%Y-%m-%d %H:%M:%S") if hasattr(val, 'strftime') else str(val)
                    elif isinstance(val, (int, float)):
                        row_dict[col] = val
                    else:
                        row_dict[col] = str(val)
                rows.append(row_dict)
            return rows
        except Exception as e:
            logger.warning(f"获取样本数据失败: {e}")
            return []

    def _is_potential_key_field(self, col_name: str, col_data: pd.Series) -> bool:
        """
        判断是否是潜在的关键字段
        """
        col_name_lower = col_name.lower()

        # 基于列名判断
        if any(keyword in col_name_lower for keyword in ["id", "编号", "key", "code"]):
            return True

        # 基于唯一性判断
        if len(col_data.dropna().unique()) == len(col_data.dropna()):
            return True

        return False

    def _convert_to_json_serializable(self, obj):
        """
        递归转换对象为可JSON序列化的格式

        Args:
            obj: 要转换的对象

        Returns:
            可JSON序列化的对象
        """
        from datetime import date, datetime

        if isinstance(obj, (datetime, date, pd.Timestamp)):
            # 如果是日期类型，检查时间部分是否为00:00:00，如果是则只返回日期部分
            if isinstance(obj, pd.Timestamp):
                if obj.hour == 0 and obj.minute == 0 and obj.second == 0:
                    return obj.strftime("%Y-%m-%d")
                else:
                    return obj.isoformat()
            elif isinstance(obj, datetime):
                if obj.hour == 0 and obj.minute == 0 and obj.second == 0:
                    return obj.strftime("%Y-%m-%d")
                else:
                    return obj.isoformat()
            elif isinstance(obj, date):
                return obj.strftime("%Y-%m-%d")
            else:
                return obj.isoformat()
        elif isinstance(obj, dict):
            return {k: self._convert_to_json_serializable(v) for k, v in obj.items()}
        elif isinstance(obj, (list, tuple)):
            return [self._convert_to_json_serializable(item) for item in obj]
        elif pd.isna(obj):
            return None
        elif isinstance(obj, (int, float, str, bool, type(None))):
            return obj
        else:
            return str(obj)

    def _process_csv_file(
        self,
        csv_file_path: str,
        table_name: str = None,
        force_reimport: bool = False,
        original_filename: str = None,
        conv_uid: str = None,
        preview_limit: int = None,
    ) -> Dict:
        """
        处理CSV文件（内存优化版本，支持大文件）
        
        Args:
            csv_file_path: CSV文件路径
            table_name: 表名
            force_reimport: 是否强制重新导入
            original_filename: 原始文件名
            conv_uid: 会话ID
            preview_limit: 预览数据行数限制
            
        Returns:
            处理结果字典
        """
        # 计算文件哈希
        content_hash = self.cache_manager.calculate_file_hash(csv_file_path)
        
        # 检查缓存
        if not force_reimport:
            cached_info = self.cache_manager.get_cached_info(content_hash)
            if cached_info and os.path.exists(cached_info["db_path"]):
                logger.info(f"CSV缓存命中: {original_filename}")
                
                # 获取预览数据
                preview_data = self.get_table_preview_data(
                    cached_info["db_path"],
                    cached_info["table_name"],
                    preview_limit,
                    original_filename
                )
                
                # 获取top_10_rows
                try:
                    import duckdb
                    conn = duckdb.connect(cached_info["db_path"], read_only=True)
                    columns_result = conn.execute(
                        f'DESCRIBE "{cached_info["table_name"]}"'
                    ).fetchall()
                    columns = [col[0] for col in columns_result]
                    rows = conn.execute(
                        f'SELECT * FROM "{cached_info["table_name"]}" LIMIT 10'
                    ).fetchall()
                    conn.close()
                    top_10_rows = [dict(zip(columns, row)) for row in rows]
                    top_10_rows = self._convert_to_json_serializable(top_10_rows)
                except Exception as e:
                    logger.warning(f"从数据库读取top_10_rows失败: {e}")
                    top_10_rows = []
                
                return {
                    "status": "cached",
                    "message": "使用缓存数据",
                    "content_hash": content_hash,
                    "db_name": cached_info["db_name"],
                    "db_path": cached_info["db_path"],
                    "table_name": cached_info["table_name"],
                    "row_count": cached_info["row_count"],
                    "column_count": cached_info["column_count"],
                    "columns_info": cached_info["columns_info"],
                    "summary_prompt": cached_info["summary_prompt"],
                    "data_schema_json": cached_info.get("data_schema_json"),
                    "id_columns": cached_info.get("id_columns", []),
                    "top_10_rows": top_10_rows,
                    "preview_data": preview_data,
                    "access_count": cached_info["access_count"],
                    "last_accessed": cached_info["last_accessed"],
                    "conv_uid": conv_uid,
                }
        
        # 没有缓存，开始处理
        logger.info(f"处理CSV文件: {original_filename}")
        
        # 获取LLM客户端
        if self.llm_client is None or self.model_name is None:
            llm_client, model_name = self._get_llm_client_and_model()
            if self.llm_client is None and llm_client is not None:
                self.llm_client = llm_client
            if self.model_name is None and model_name is not None:
                self.model_name = model_name
        
        # 生成表名
        if table_name is None:
            base_name = Path(original_filename).stem
            base_name = "".join(
                c if c.isalnum() or c == "_" else "_" for c in base_name
            )
            if base_name and base_name[0].isdigit():
                base_name = f"tbl_{base_name}"
            if not base_name or len(base_name) < 2:
                base_name = f"csv_table_{content_hash[:8]}"
            table_name = base_name
        
        db_name = f"csv_{content_hash[:8]}"
        db_filename = f"{db_name}.duckdb"
        db_path = str(self.db_storage_dir / db_filename)
        
        print(f"[DEBUG] CSV数据库路径: {db_path}, 表名: {table_name}")
        
        # 使用DuckDB直接读取CSV（内存优化）
        actual_row_count, columns = self.process_csv_to_duckdb(
            csv_file_path, db_path, table_name
        )
        
        # 从DuckDB加载样本数据用于分析（只加载1000行）
        import duckdb
        conn = duckdb.connect(db_path)
        df_sample = conn.execute(f'SELECT * FROM "{table_name}" LIMIT 1000').df()
        conn.close()
        
        print(f"[DEBUG] 加载样本数据: {len(df_sample)} 行用于分析")
        
        # 数据清洗（仅对样本）
        df_sample = self._remove_empty_columns(df_sample)
        df_sample = self._remove_duplicate_columns(df_sample)
        df_sample = self._format_date_columns(df_sample)
        
        # 识别ID列
        id_columns = []
        try:
            id_columns = self._detect_id_columns_with_llm(df_sample, table_name)
            if id_columns:
                logger.info(f"ID列: {id_columns}")
                print(f"[DEBUG] 检测到ID列: {id_columns}")
        except Exception as e:
            logger.warning(f"识别ID列失败: {e}")
        
        # 对样本数据进行类型转换
        df_sample = self._convert_id_columns_to_string(df_sample, id_columns)
        df_sample = self._convert_column_types(df_sample, id_columns)
        df_sample = self._format_numeric_columns(df_sample, id_columns)
        
        # 清理列名
        df_sample.columns = [
            str(col)
            .replace(" ", "")
            .replace("\u00a0", "")
            .replace("\n", "")
            .replace("\r", "")
            .replace("\t", "")
            for col in df_sample.columns
        ]
        
        # 去重列名
        final_columns = []
        seen_columns = {}
        for col in df_sample.columns:
            if col in seen_columns:
                seen_columns[col] += 1
                final_columns.append(f"{col}_{seen_columns[col]}")
            else:
                seen_columns[col] = 0
                final_columns.append(col)
        df_sample.columns = final_columns
        
        print(f"[DEBUG] 样本数据处理完成，列数: {len(df_sample.columns)}")
        
        # 获取列信息
        conn = duckdb.connect(db_path)
        try:
            columns_result = conn.execute(f'DESCRIBE "{table_name}"').fetchall()
            columns_info = [
                {"name": col[0], "type": col[1], "dtype": col[1]}
                for col in columns_result
            ]
        finally:
            conn.close()
        
        # 生成schema
        try:
            schema_understanding_json = self._generate_schema_understanding_with_llm(
                df_sample, table_name, id_columns
            )
        except Exception as llm_e:
            logger.warning(f"LLM生成schema失败，使用基础schema: {llm_e}")
            schema_understanding_json = self._generate_basic_schema_json(
                df_sample, table_name, id_columns
            )
        
        summary_prompt = self._format_schema_as_prompt(
            schema_understanding_json, df_sample, table_name
        )
        
        # 保存缓存
        self.cache_manager.save_cache_info(
            content_hash=content_hash,
            original_filename=original_filename,
            table_name=table_name,
            db_name=db_name,
            db_path=db_path,
            df=df_sample,
            summary_prompt=summary_prompt,
            data_schema_json=schema_understanding_json,
            id_columns=id_columns,
        )
        
        # 注册到DB-GPT
        self._register_to_dbgpt(db_name, db_path, table_name)
        
        # 获取top_10_rows
        top_10_rows_raw = df_sample.head(10).values.tolist()
        top_10_rows = self._convert_to_json_serializable(top_10_rows_raw)
        
        # 获取预览数据
        preview_data = self.get_table_preview_data(
            db_path, table_name, preview_limit, original_filename
        )
        
        return {
            "status": "imported",
            "message": "成功导入CSV数据",
            "content_hash": content_hash,
            "db_name": db_name,
            "db_path": db_path,
            "table_name": table_name,
            "row_count": actual_row_count,
            "column_count": len(df_sample.columns),
            "columns_info": columns_info,
            "summary_prompt": summary_prompt,
            "data_schema_json": schema_understanding_json,
            "id_columns": id_columns,
            "top_10_rows": top_10_rows,
            "preview_data": preview_data,
            "conv_uid": conv_uid,
        }

    def _register_to_dbgpt(self, db_name: str, db_path: str, table_name: str):
        """注册到 DB-GPT 数据源管理器"""
        try:
            logger.info(
                f"SQLite数据库已创建: {db_name}, 路径: {db_path}, 表名: {table_name}"
            )
            return
        except Exception as e:
            logger.warning(f"注册到 DB-GPT 失败: {e}")

    def update_summary_prompt(self, content_hash: str, summary_prompt: str):
        """
        更新数据理解提示词

        Args:
            content_hash: 内容哈希值
            summary_prompt: 新的数据理解提示词
        """
        self.cache_manager.update_summary_prompt(content_hash, summary_prompt)

    def get_excel_info(self, content_hash: str) -> Optional[Dict]:
        """
        获取 Excel 信息

        Args:
            content_hash: 内容哈希值

        Returns:
            Excel 信息字典
        """
        return self.cache_manager.get_cached_info(content_hash)

    def get_tables_info_for_selection(self, file_hash: str) -> List[Dict]:
        """
        获取多表信息用于表选择
        
        Args:
            file_hash: 文件哈希值
            
        Returns:
            表信息列表，每个元素包含表选择所需的信息
        """
        tables = self.cache_manager.get_tables_by_file_hash(file_hash)
        
        result = []
        for table in tables:
            result.append({
                "table_name": table.get("table_name"),
                "sheet_name": table.get("sheet_name"),
                "table_hash": table.get("table_hash"),
                "create_table_sql": table.get("create_table_sql"),
                "data_schema_json": table.get("data_schema_json"),
                "row_count": table.get("row_count"),
                "column_count": table.get("column_count"),
                "db_path": table.get("db_path"),
            })
        
        return result

    def get_table_schema_by_name(
        self, file_hash: str, table_name: str
    ) -> Optional[Dict]:
        """
        根据文件哈希和表名获取单个表的完整信息
        
        Args:
            file_hash: 文件哈希值
            table_name: 表名
            
        Returns:
            表的完整信息字典
        """
        tables = self.cache_manager.get_tables_by_file_hash(file_hash)
        
        for table in tables:
            if table.get("table_name") == table_name:
                return table
        
        return None

    def get_combined_schema_for_tables(
        self, file_hash: str, table_names: List[str]
    ) -> Dict:
        """
        获取多个表的组合Schema信息，用于query改写
        
        Args:
            file_hash: 文件哈希值
            table_names: 要组合的表名列表
            
        Returns:
            组合后的Schema信息
        """
        tables = self.cache_manager.get_tables_by_file_hash(file_hash)
        
        selected_tables = []
        for table in tables:
            if table.get("table_name") in table_names:
                selected_tables.append(table)
        
        if not selected_tables:
            return {}
        
        # 如果只有一个表，直接返回其schema
        if len(selected_tables) == 1:
            table = selected_tables[0]
            return {
                "table_name": table.get("table_name"),
                "data_schema_json": table.get("data_schema_json"),
                "create_table_sql": table.get("create_table_sql"),
                "summary_prompt": table.get("summary_prompt"),
                "db_path": table.get("db_path"),
            }
        
        # 多个表的情况，组合schema
        combined_columns = []
        table_descriptions = []
        create_sqls = []
        
        for table in selected_tables:
            table_name = table.get("table_name")
            schema_json = table.get("data_schema_json")
            
            if schema_json:
                try:
                    schema = json.loads(schema_json) if isinstance(schema_json, str) else schema_json
                    
                    # 为每个列添加表名前缀
                    for col in schema.get("columns", []):
                        col_copy = col.copy()
                        col_copy["table_name"] = table_name
                        col_copy["full_column_name"] = f"{table_name}.{col.get('column_name', '')}"
                        combined_columns.append(col_copy)
                    
                    if schema.get("table_description"):
                        table_descriptions.append(f"【{table_name}】: {schema['table_description']}")
                except Exception as e:
                    logger.warning(f"解析表 {table_name} 的schema失败: {e}")
            
            if table.get("create_table_sql"):
                create_sqls.append(table.get("create_table_sql"))
        
        combined_schema = {
            "tables": [t.get("table_name") for t in selected_tables],
            "table_descriptions": "\n".join(table_descriptions),
            "columns": combined_columns,
            "create_table_sqls": "\n\n".join(create_sqls),
        }
        
        return {
            "is_multi_table": True,
            "table_names": [t.get("table_name") for t in selected_tables],
            "data_schema_json": json.dumps(combined_schema, ensure_ascii=False),
            "create_table_sql": "\n\n".join(create_sqls),
            "summary_prompt": "\n".join(table_descriptions),
            "db_path": selected_tables[0].get("db_path"),  # 所有表在同一个数据库
        }
